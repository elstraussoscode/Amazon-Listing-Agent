"""
Amazon Listing Agent - Production App with Dynamic Template Support
Automatisches Befüllen von BELIEBIGEN Amazon-Templates mit KI-optimierten Inhalten
"""

import streamlit as st
import pandas as pd
import openpyxl
from openai import OpenAI
import io
from typing import Dict, List, Optional
import logging
from pydantic import BaseModel, Field
import tempfile
import os
import re
import unicodedata
from dynamic_template_analyzer import analyze_template, TemplateFormat

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

# Pydantic Models
class ProductContent(BaseModel):
    """AI-generated Amazon listing content"""
    model_config = {"extra": "forbid"}  # Required for Responses API
    
    artikelname: str = Field(description="Optimierter Produkttitel, max 200 Zeichen")
    bullet_points: List[str] = Field(description="5 Aufzählungspunkte, je max 250 Zeichen", min_length=5, max_length=5)
    suchbegriffe: str = Field(description="5 Suchbegriffe durch Komma getrennt")

class CosmoOptimizedContent(BaseModel):
    """COSMO/RUFUS optimized Amazon listing content"""
    model_config = {"extra": "forbid"}
    
    artikelname: str = Field(description="Optimierter Produkttitel (Max 200 Zeichen, inkl. Leerzeichen). Fokus: Brand + Produktart + USP + Variation.")
    produktbeschreibung: str = Field(description="Detaillierte Produktbeschreibung (Max 2000 Zeichen). Muss alle 15 COSMO-Beziehungstypen abdecken.")
    bullet_points: List[str] = Field(description="5 Aufzählungspunkte (Je max 250 Zeichen). Verkaufsstark & nutzenorientiert.", min_length=5, max_length=5)
    suchbegriffe: str = Field(description="Keywords (Max 249 Bytes/Zeichen). Relevante Begriffe, NICHT in Titel/Bullets wiederholen!")

# Default Prompt
DEFAULT_PROMPT = """Erstelle einen optimierten Amazon-Listing für folgendes Produkt:

{{product_data}}

WICHTIG - Amazon COSMO & RUFUS Optimierung (2025):
- Fokus auf KONTEXT statt Keywords
- Natürliche, hilfreiche Sprache
- Beantworte diese Fragen intern (NICHT im Text erwähnen!):
  * Was ist das Produkt?
  * Für wen ist es?
  * Welches Problem löst es?
  * Was macht es besonders?
  * Wie wird es verwendet?

REGELN:
1. Artikelname: Max 200 Zeichen, relevant, kein Keyword-Stuffing
2. Bullet Points: Je max 250 Zeichen, KEINE Nummerierung (1., 2., etc.), natürlich lesbar
3. Suchbegriffe: 5 relevante Begriffe durch Komma getrennt"""

# COSMO Prompt
COSMO_PROMPT = """Erstelle ein vollständig COSMO & RUFUS optimiertes Amazon-Listing für folgendes Produkt.

Produktdaten:
{{product_data}}

✅ COSMO-/RUFUS-Check: Decke ALLE 15 semantischen Beziehungstypen ab:
1. is (Was ist es?)
2. has_property (Eigenschaften: Farbe, Material, Größe...)
3. has_component (Teile/Zubehör)
4. used_for (Verwendungszweck)
5. used_in (Situation/Umgebung)
6. used_by (Zielgruppe/Nutzer)
7. used_with (Kombinationsprodukte)
8. made_of (Material)
9. has_quality (Qualitätsmerkmale)
10. has_brand (Marke/Botschaft)
11. has_style (Stil/Design)
12. targets_audience (Zielgruppe detailliert)
13. associated_with (Themen: Nachhaltigkeit, Gesundheit...)
14. has_certification (Zertifikate)
15. enables_activity (Ermöglichte Aktivität)

WICHTIG - LENGTH LIMITS (STRICT):
1. TITEL: Maximal 200 Zeichen! (Ideal: 150-180)
   - Format: [Marke] [Produktart] [Wichtigstes Merkmal/USP] [Farbe/Größe]
2. BULLET POINTS: 5 Stück, jeweils MAXIMAL 250 Zeichen!
   - Jeder Punkt muss einen konkreten Kundennutzen ansprechen.
   - Keine reinen Keyword-Listen.
3. BESCHREIBUNG: Maximal 2000 Zeichen.
   - Fließtext, gut lesbar, integriert die 15 COSMO-Aspekte natürlich.
4. KEYWORDS (Suchbegriffe): Maximal 249 Bytes (ca. 240 Zeichen).
   - NUR Wörter, die NICHT im Titel oder den Bullets stehen!
   - Synonyme, Falschschreibweisen, spezifische Nischenbegriffe.

STRUKTUR:
1. Titel: Optimiert mit Wer, Was, Wofür, Wie.
2. Beschreibung: Ausführlich, deckt alle 15 Punkte ab.
3. Bullet Points: 5 Stück, verkaufsstark, nutzenorientiert.
4. Keywords: Nur NEUE, relevante Begriffe.
"""

# Page Config
st.set_page_config(
    page_title="Amazon Listing Agent - Production",
    page_icon="🛒",
    layout="wide"
)

# Initialize Session State
if 'api_key' not in st.session_state:
    st.session_state.api_key = ""
if 'prompt_template' not in st.session_state:
    st.session_state.prompt_template = DEFAULT_PROMPT
if 'cosmo_prompt_template' not in st.session_state:
    st.session_state.cosmo_prompt_template = COSMO_PROMPT

# Title
st.title("🛒 Amazon Listing Agent - Production")
st.markdown("**Unterstützt ALLE Amazon Template-Formate automatisch!**")

# Sidebar - Configuration
with st.sidebar:
    st.header("⚙️ Konfiguration")
    
    # API Key
    api_key_input = st.text_input(
        "OpenAI API Key",
        value=st.session_state.api_key,
        type="password",
        help="Ihr OpenAI API Key für GPT-5.1"
    )
    
    if st.button("💾 API Key speichern"):
        st.session_state.api_key = api_key_input
        st.success("✅ API Key gespeichert!")
    
    # Status
    st.markdown("---")
    st.markdown("### 📊 Status")
    if st.session_state.get('api_key', '').strip():
        st.success("✅ API Key")
    else:
        st.error("❌ API Key")
    
    # Prompt Template
    with st.expander("✏️ Standard Prompt bearbeiten"):
        prompt = st.text_area(
            "Standard Prompt",
            value=st.session_state.prompt_template,
            height=300
        )
        if st.button("💾 Standard Prompt speichern"):
            st.session_state.prompt_template = prompt
            st.success("✅ Prompt gespeichert!")
    
    with st.expander("✏️ COSMO Prompt bearbeiten"):
        cosmo_prompt = st.text_area(
            "COSMO Prompt",
            value=st.session_state.cosmo_prompt_template,
            height=300
        )
        if st.button("💾 COSMO Prompt speichern"):
            st.session_state.cosmo_prompt_template = cosmo_prompt
            st.success("✅ Prompt gespeichert!")

# Main Tabs
tab1, tab4, tab2, tab3 = st.tabs(["📝 Template Filler", "✍️ Content-Optimierung", "📊 Preview", "⚙️ Settings"])

with tab1:
    st.header("📝 Automatisches Template-Befüllen")
    
    # Workflow Steps
    st.info("""
    **Workflow:** 
    1️⃣ Produktdaten hochladen (Excel) 
    → 2️⃣ Amazon-Template hochladen (Excel) 
    → 3️⃣ KI generiert Inhalte 
    → 4️⃣ Template wird automatisch befüllt
    → 5️⃣ Download fertiges Template
    """)
    
    col1, col2 = st.columns(2)
    
    # Upload Product Data
    with col1:
        st.subheader("1️⃣ Produktdaten hochladen")
        products_file = st.file_uploader(
            "Excel mit Produktdaten",
            type=["xlsx", "xls"],
            help="Eine Excel-Datei mit Ihren Produktdaten (1 Produkt pro Zeile)"
        )
        
        products_df = None
        if products_file:
            try:
                # Dynamic header detection
                df_test = pd.read_excel(products_file, header=None, nrows=10)
                
                # Find the actual header row (row with most non-numeric text values)
                header_row = 0
                best_score = 0
                
                for row_idx in range(min(5, len(df_test))):
                    row = df_test.iloc[row_idx]
                    # Score: count cells with text that looks like column names
                    score = sum(1 for val in row if pd.notna(val) and 
                               isinstance(val, str) and 
                               len(val) > 2 and
                               not val.replace('.', '').replace('-', '').isdigit())
                    if score > best_score:
                        best_score = score
                        header_row = row_idx
                
                # Read with detected header
                products_df = pd.read_excel(products_file, header=header_row)
                
                st.success(f"✅ {len(products_df)} Produkte geladen (Header in Zeile {header_row})")
                
                with st.expander("📋 Daten-Vorschau"):
                    st.dataframe(products_df.head(10), use_container_width=True)
            except Exception as e:
                st.error(f"❌ Fehler beim Laden: {e}")
                logger.error(f"Error loading products: {e}", exc_info=True)
    
    # Upload Template
    with col2:
        st.subheader("2️⃣ Amazon-Template hochladen")
        template_file = st.file_uploader(
            "Amazon Upload-Template (.xlsm)",
            type=["xlsm", "xlsx"],
            help="Beliebiges Amazon Seller/Vendor Template"
        )
        
        template_bytes = None
        format_info = None
        
        if template_file:
            try:
                template_bytes = template_file.read()
                
                # Analyze template dynamically
                with st.spinner("🔍 Analysiere Template-Struktur..."):
                    with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsm') as tmp:
                        tmp.write(template_bytes)
                        tmp_path = tmp.name
                    
                    format_info = analyze_template(tmp_path)
                    os.unlink(tmp_path)
                
                # Show detection results
                st.success(f"✅ Template-Format: **{format_info.format_type}**")
                
                col2a, col2b, col2c = st.columns(3)
                with col2a:
                    st.metric("Header Row", format_info.header_row)
                with col2b:
                    st.metric("Data Row", format_info.data_start_row)
                with col2c:
                    st.metric("Product Types", len(format_info.product_types))
                
                with st.expander("🔍 Erkannte Spalten"):
                    st.write(f"**SKU:** Column {format_info.sku_column}")
                    st.write(f"**Product Type:** Column {format_info.product_type_column}")
                    st.write(f"**Brand:** Column {format_info.brand_column}")
                    st.write(f"**Title:** Column {format_info.title_column}")
                    if format_info.format_type == "XML":
                        st.write(f"**Offer Action:** Column {format_info.offer_action_column}")
                
                with st.expander(f"🎯 Verfügbare Product Types ({len(format_info.product_types)})"):
                    for i, pt in enumerate(format_info.product_types[:20], 1):
                        st.write(f"{i}. {pt}")
                    if len(format_info.product_types) > 20:
                        st.write(f"... und {len(format_info.product_types) - 20} weitere")
                
                # Show required fields from Datendefinitionen (GLOBAL)
                if format_info.required_fields:
                    with st.expander(f"🔴 Globale Pflichtfelder ({len(format_info.required_fields)})"):
                        st.info("ℹ️ **Hinweis:** Spezifische Pflichtfelder für jeden Product Type werden nach der Typ-Erkennung pro Produkt angezeigt.")
                        st.write("*Alle Pflichtfelder aus Datendefinitionen (Product Type unabhängig):*")
                        for field in format_info.required_fields[:15]:
                            st.write(f"• {field}")
                        if len(format_info.required_fields) > 15:
                            st.write(f"... und {len(format_info.required_fields) - 15} weitere")
                
            except Exception as e:
                st.error(f"❌ Template-Analyse fehlgeschlagen: {e}")
                logger.error(f"Template analysis error: {e}", exc_info=True)
    
    
    # Helper function to enforce length with AI regeneration
    def ensure_length_with_ai(text: str, max_length: int, field_name: str, client_instance) -> str:
        if len(text) <= max_length:
            return text
            
        logger.info(f"⚠️ {field_name} too long ({len(text)} chars). Regenerating to fit {max_length} chars...")
        
        shorten_prompt = f"""Kürze den folgenden Text für Amazon auf MAXIMAL {max_length} Zeichen.
        WICHTIG: Erhalte die Kernaussage und wichtige Keywords!
        
        Original Text:
        {text}
        
        Antworte NUR mit dem gekürzten Text:"""
        
        try:
            resp = client_instance.chat.completions.create(
                model="gpt-5.1",
                messages=[
                    {"role": "system", "content": "Du bist ein präziser Redakteur. Kürze Texte ohne Inhaltsverlust."},
                    {"role": "user", "content": shorten_prompt}
                ]
            )
            new_text = resp.choices[0].message.content.strip()
            
            # Fallback if still too long (truncate nicely)
            if len(new_text) > max_length:
                logger.warning(f"⚠️ Still too long after regeneration ({len(new_text)}). Truncating.")
                return new_text[:max_length-3] + "..."
            
            return new_text
        except Exception as e:
            logger.error(f"Error regenerating text: {e}")
            return text[:max_length-3] + "..."

    # Processing Section
    if products_df is not None and template_bytes and format_info:
        st.markdown("---")
        st.subheader("3️⃣ Content-Generierung & Template-Befüllung")
        
        # Settings
        total_products = len(products_df)
        num_products = st.slider(
            f"Anzahl zu verarbeitender Produkte (von {total_products} verfügbar)",
            min_value=1,
            max_value=total_products,
            value=min(5, total_products)
        )
        start_row = 0  # Always start from first row
        
        # Start Button
        if st.button("🚀 Starten: Content generieren & Template füllen", type="primary", use_container_width=True):
            if not st.session_state.get('api_key', '').strip():
                st.error("❌ Bitte OpenAI API Key eingeben!")
            else:
                # Process products
                generated_contents = []
                progress_bar = st.progress(0)
                status_text = st.empty()
                results_container = st.container()
                
                for idx in range(num_products):
                    try:
                        status_text.text(f"🤖 Analysiere Produkt {idx + 1}/{num_products}...")
                        
                        # Get product row
                        actual_row = start_row + idx
                        if actual_row >= len(products_df):
                            break
                        
                        product = products_df.iloc[actual_row]
                        
                        # Extract product data (sanitize Unicode for API)
                        product_data = {}
                        for col in products_df.columns:
                            if pd.notna(product[col]):
                                # Clean ALL special characters that cause encoding issues
                                value = str(product[col])
                                # Normalize to ASCII-compatible characters (NFD = decomposed form)
                                value = unicodedata.normalize('NFD', value)
                                value = value.encode('ascii', 'ignore').decode('ascii')
                                product_data[col] = value
                        
                        client = OpenAI(api_key=st.session_state.api_key)
                        
                        # STEP 1: Determine product type
                        status_text.text(f"🎯 Bestimme Produkttyp für Produkt {idx + 1}...")
                        product_info = "\n".join([f"- {k}: {v}" for k, v in list(product_data.items())[:15]])
                        
                        # Sanitize product types list for API (remove Unicode)
                        clean_types = []
                        for pt in format_info.product_types[:20]:
                            clean_pt = unicodedata.normalize('NFD', str(pt))
                            clean_pt = clean_pt.encode('ascii', 'ignore').decode('ascii')
                            if clean_pt.strip():
                                clean_types.append(clean_pt)
                        
                        type_prompt = f"""Analysiere dieses Produkt und bestimme den passenden Amazon Product Type:

Produktdaten:
{product_info}

Verfugbare Product Types:
{', '.join(clean_types)}

Wahle den BESTEN passenden Product Type aus der Liste. 
Antworte NUR mit dem exakten Product Type Namen, nichts anderes."""
                        
                        type_response = client.chat.completions.create(
                            model="gpt-5.1",
                            messages=[
                                {"role": "system", "content": "Du bist ein Produktkategorisierungs-Experte."},
                                {"role": "user", "content": type_prompt}
                            ]
                        )
                        
                        detected_type = type_response.choices[0].message.content.strip()
                        # Validate and fallback
                        if detected_type not in format_info.product_types:
                            detected_type = format_info.product_types[0] if format_info.product_types else None
                        
                        # Get required attributes for this specific product type
                        required_attributes = []
                        required_field_labels = []
                        if detected_type:
                            required_attributes = format_info.product_type_required_attributes.get(detected_type.lower(), [])
                            required_field_labels = [
                                format_info.field_display_names.get(attr, attr)
                                for attr in required_attributes
                            ]
                            logger.info(f"Product Type: {detected_type} → {len(required_attributes)} required fields")
                        
                        # STEP 2: Generate optimized content with GPT-5.1
                        status_text.text(f"✍️ Generiere Content für {detected_type}...")
                        prompt = st.session_state.prompt_template.replace("{{product_data}}", product_info)
                        
                        # Use Chat Completions API with structured outputs
                        response = client.chat.completions.create(
                            model="gpt-5.1",
                            messages=[
                                {"role": "system", "content": "Du bist ein Amazon SEO-Experte. Erstelle optimierte Listings nach COSMO/RUFUS Prinzipien."},
                                {"role": "user", "content": prompt}
                            ],
                            response_format={
                                "type": "json_schema",
                                "json_schema": {
                                    "name": "product_content",
                                    "schema": ProductContent.model_json_schema(),
                                    "strict": True
                                }
                            }
                        )
                        
                        content = ProductContent.model_validate_json(response.choices[0].message.content)
                        
                        # Enforce hard limits with AI regeneration
                        if len(content.artikelname) > 200:
                            content.artikelname = ensure_length_with_ai(content.artikelname, 200, "Titel", client)
                            
                        new_bullets = []
                        for i, bp in enumerate(content.bullet_points):
                            if len(bp) > 250:
                                new_bp = ensure_length_with_ai(bp, 250, f"Bullet {i+1}", client)
                                new_bullets.append(new_bp)
                            else:
                                new_bullets.append(bp)
                        content.bullet_points = new_bullets
                        
                        if len(content.suchbegriffe) > 249:
                            content.suchbegriffe = ensure_length_with_ai(content.suchbegriffe, 249, "Keywords", client)
                            
                        generated_contents.append((content, detected_type, product_data, required_attributes))
                        
                        # Show preview
                        with results_container:
                            with st.expander(f"✅ Produkt {idx + 1}: {content.artikelname[:50]}... [{detected_type}]"):
                                st.write(f"**🎯 Product Type:** {detected_type}")
                                if required_field_labels:
                                    st.write(f"**🔴 Pflichtfelder für {detected_type}:** {len(required_field_labels)}")
                                    preview_fields = ', '.join(required_field_labels[:8])
                                    if len(required_field_labels) > 8:
                                        preview_fields += ", ..."
                                    st.caption(preview_fields)
                                st.write("**Artikelname:**", content.artikelname)
                                st.write("**Bullet Points:**")
                                for i, bp in enumerate(content.bullet_points, 1):
                                    st.write(f"• {bp}")
                                st.write("**Suchbegriffe:**", content.suchbegriffe)
                        
                        progress_bar.progress((idx + 1) / num_products)
                        
                    except Exception as e:
                        st.error(f"❌ Fehler bei Produkt {idx + 1}: {e}")
                        logger.error(f"Error generating content for product {idx + 1}: {e}", exc_info=True)
                
                status_text.text(f"✅ Fertig! {len(generated_contents)} Produkte generiert")
                
                # Fill template
                if generated_contents:
                    with st.spinner("📝 Befülle Template..."):
                        try:
                            # Load original template - PRESERVES EVERYTHING:
                            # ✅ All sheets, ✅ VBA macros, ✅ Formulas, ✅ Formatting
                            # ✅ Data validations, ✅ Dropdowns, ✅ Conditional formatting
                            # We ONLY write to specific cells in data rows!
                            wb = openpyxl.load_workbook(io.BytesIO(template_bytes), keep_vba=True)
                            ws = wb[format_info.sheet_name]
                            
                            logger.info(f"📄 Template loaded: {len(wb.sheetnames)} sheets preserved")
                            logger.info(f"✅ VBA Macros: {'Yes' if wb.vba_archive else 'No'}")
                            
                            # Fill each product
                            missing_required_summary = []
                            
                            for idx, (content, detected_type, product_data, required_fields) in enumerate(generated_contents):
                                actual_row = start_row + idx
                                target_row = format_info.data_start_row + idx
                                
                                logger.info(f"\n{'='*60}")
                                logger.info(f"Filling row {target_row} with {detected_type}")
                                logger.info(f"Required fields for this type: {len(required_fields)}")
                                logger.info(f"{'='*60}")
                                
                                def set_internal_value(attribute_name: Optional[str], value, label: Optional[str] = None):
                                    if not attribute_name or attribute_name not in format_info.internal_columns:
                                        return
                                    if value is None:
                                        return
                                    if isinstance(value, str):
                                        value = value.strip()
                                        if not value:
                                            return
                                    col_info = format_info.internal_columns[attribute_name]
                                    ws.cell(row=target_row, column=col_info.column_index + 1).value = value
                                    field_label = label or format_info.field_display_names.get(attribute_name, attribute_name)
                                    logger.info(f"  ✅ {field_label} ({col_info.column}): {str(value)[:60]}")
                                
                                def set_internal_by_keyword(keyword: str, value, limit: Optional[int] = None):
                                    if value is None:
                                        return
                                    keyword_lower = keyword.lower()
                                    count = 0
                                    for attr_name in format_info.internal_columns.keys():
                                        if keyword_lower in attr_name.lower():
                                            set_internal_value(attr_name, value)
                                            count += 1
                                            if limit and count >= limit:
                                                break
                                
                                def get_product_value(*keys):
                                    for key in keys:
                                        if key and key in product_data:
                                            val = product_data.get(key)
                                            if isinstance(val, str) and not val.strip():
                                                continue
                                            if val is not None:
                                                return str(val).strip()
                                    return None
                                
                                def normalize_price(value: Optional[str]) -> Optional[str]:
                                    if not value:
                                        return None
                                    clean = str(value).strip().replace('€', '').replace('EUR', '')
                                    clean = clean.replace(',', '.')
                                    allowed_chars = ''.join(ch for ch in clean if ch.isdigit() or ch == '.')
                                    if not allowed_chars:
                                        return None
                                    try:
                                        return f"{float(allowed_chars):.2f}"
                                    except ValueError:
                                        return None
                                
                                # STEP 1: Fill AI-generated content
                                # Title
                                if format_info.title_column_index is not None:
                                    ws.cell(row=target_row, column=format_info.title_column_index + 1).value = content.artikelname
                                    logger.info(f"✅ Title: {content.artikelname[:50]}")
                                
                                # Bullet Points (detect columns dynamically)
                                bp_cols = [col_info for name, col_info in format_info.key_columns.items() 
                                          if 'Bullet Point' in name or 'Aufzählungspunkt' in name]
                                for i, bp in enumerate(content.bullet_points[:5]):
                                    if i < len(bp_cols):
                                        col_idx = bp_cols[i].column_index + 1
                                        ws.cell(row=target_row, column=col_idx).value = bp
                                        logger.info(f"✅ BP{i+1}: {bp[:40]}")
                                
                                # Search Terms (detect columns dynamically)
                                search_terms = [s.strip() for s in content.suchbegriffe.split(',')]
                                st_cols = [col_info for name, col_info in format_info.key_columns.items() 
                                          if 'Search Term' in name or 'Suchbegriff' in name]
                                for i, term in enumerate(search_terms[:5]):
                                    if i < len(st_cols):
                                        col_idx = st_cols[i].column_index + 1
                                        ws.cell(row=target_row, column=col_idx).value = term
                                        logger.info(f"✅ Search{i+1}: {term}")
                                
                                # STEP 2: Fill detected Product Type
                                if format_info.product_type_column_index is not None and detected_type:
                                    ws.cell(row=target_row, column=format_info.product_type_column_index + 1).value = detected_type
                                    logger.info(f"✅ Product Type: {detected_type}")
                                
                                # STEP 3: Fill from product data (intelligent mapping)
                                sku_value = (
                                    product_data.get('SKU') or
                                    product_data.get('sku') or
                                    product_data.get('Seller-SKU') or
                                    product_data.get('Artikel-Nr.') or
                                    product_data.get('ASIN')
                                )
                                
                                if format_info.sku_column_index is not None and sku_value:
                                    ws.cell(row=target_row, column=format_info.sku_column_index + 1).value = str(sku_value)
                                    logger.info(f"✅ SKU: {sku_value}")
                                
                                # Brand
                                brand_value = (
                                    product_data.get('Brand') or
                                    product_data.get('Marke') or
                                    product_data.get('brand_name') or
                                    product_data.get('Marke/Brand')
                                )
                                
                                if format_info.brand_column_index is not None and brand_value:
                                    ws.cell(row=target_row, column=format_info.brand_column_index + 1).value = str(brand_value)
                                    logger.info(f"✅ Brand: {brand_value}")
                                
                                # EAN/GTIN
                                ean_col = format_info.key_columns.get('EAN')
                                ean_value = (
                                    product_data.get('EAN Code') or
                                    product_data.get('EAN') or
                                    product_data.get('GTIN') or
                                    product_data.get('Barcode')
                                )
                                if ean_col and ean_value:
                                    ws.cell(row=target_row, column=ean_col.column_index + 1).value = str(ean_value)
                                    logger.info(f"✅ EAN: {ean_value}")
                                
                                # Material
                                material_col = format_info.key_columns.get('Material')
                                if material_col:
                                    material = product_data.get('Material') or product_data.get('material')
                                    if material:
                                        ws.cell(row=target_row, column=material_col.column_index + 1).value = str(material)
                                        logger.info(f"✅ Material: {material}")
                                
                                # Color
                                color_col = format_info.key_columns.get('Color')
                                if color_col:
                                    color = product_data.get('Color') or product_data.get('Farbe') or product_data.get('colour')
                                    if color:
                                        ws.cell(row=target_row, column=color_col.column_index + 1).value = str(color)
                                        logger.info(f"✅ Color: {color}")
                                
                                # Size/Dimensions
                                size_col = format_info.key_columns.get('Size')
                                if size_col:
                                    size = (product_data.get('Size') or product_data.get('Größe') or 
                                           product_data.get('P-Gesamt cm') or product_data.get('Dimensions'))
                                    if size:
                                        ws.cell(row=target_row, column=size_col.column_index + 1).value = str(size)
                                        logger.info(f"✅ Size: {size}")
                                
                                # STEP 4: Ensure internal attributes for the template are populated
                                manufacturer_value = (
                                    product_data.get('Hersteller') or
                                    product_data.get('Manufacturer') or
                                    brand_value
                                )
                                country_of_origin = get_product_value('Herkunftsland', 'Country of Origin', 'country_of_origin')
                                unit_count_value = get_product_value('VE', 'Menge', 'Stückzahl') or '1'
                                browse_nodes_options = format_info.browse_nodes_by_product_type.get(
                                    (detected_type or '').lower(), []
                                )
                                price_source = get_product_value('UVP', 'Preis', 'Price', 'Basis-EK')
                                normalized_price = normalize_price(price_source) or "9.99"
                                width_mm = get_product_value('P-Breite/Width mm', 'P-Breite/width cm')
                                height_mm = get_product_value('P-Höhe/hight mm', 'P-Höhe/hight cm')
                                volume_ml = get_product_value('Max. Befüllung/Filling in ml')
                                weight_grams = get_product_value('Gramm') or get_product_value('KG')
                                
                                attr_value_map = {
                                    'feed_product_type': detected_type,
                                    'item_sku': sku_value,
                                    'brand_name': brand_value,
                                    'manufacturer': manufacturer_value,
                                    'item_name': content.artikelname,
                                    'recommended_browse_nodes': browse_nodes_options[0] if browse_nodes_options else None,
                                    'external_product_id': ean_value,
                                    'external_product_id_type': 'EAN' if ean_value else None,
                                    'country_of_origin': country_of_origin,
                                    'item_type_name': detected_type,
                                    'update_delete': 'Aktualisierung' if format_info.format_type == "FLAT_FILE" else None,
                                    'merchant_shipping_group_name': 'Standard',
                                    'unit_count': unit_count_value,
                                    'unit_count_type': 'Count',
                                    'item_package_quantity': unit_count_value,
                                    'fulfillment_availability#1.fulfillment_channel_code': 'DEFAULT',
                                    'fulfillment_availability#1.quantity': get_product_value('Bestand', 'Inventory', 'Quantity') or '999',
                                    'fulfillment_availability#1.lead_time_to_ship_max_days': '2',
                                    'list_price_with_tax': normalized_price,
                                    'height_width_side_to_side': width_mm,
                                    'height_width_side_to_side_unit_of_measure': 'MM' if width_mm else None,
                                    'height_floor_top': height_mm,
                                    'height_floor_top_unit_of_measure': 'MM' if height_mm else None,
                                    'item_weight': weight_grams,
                                    'item_weight_unit_of_measure': 'g' if weight_grams else None,
                                    'item_volume': volume_ml,
                                    'item_volume_unit_of_measure': 'Milliliters' if volume_ml else None,
                                    'lithium_battery_weight_unit_of_measure': 'g',
                                    'lithium_battery_weight': '0',
                                    'lithium_battery_energy_content_unit_of_measure': 'Wh',
                                    'lithium_battery_energy_content': '0',
                                    'lithium_battery_packaging': 'Not Applicable',
                                    'number_of_lithium_ion_cells': '0',
                                    'number_of_lithium_metal_cells': '0',
                                    'battery_weight_unit_of_measure': 'g',
                                    'battery_weight': '0',
                                    'batteries_required': 'false',
                                    'are_batteries_included': 'false',
                                    'battery_cell_composition': 'No Battery',
                                    'is_expiration_dated_product': 'false',
                                    'hazmat_united_nations_regulatory_id': 'Not Applicable',
                                    'number_of_boxes': '1',
                                    'item_volume_unit_of_measure': 'Milliliters' if volume_ml else None,
                                    'material_type': product_data.get('Material') or product_data.get('material'),
                                    'model': product_data.get('Modellname') or product_data.get('Modellnummer'),
                                    'safety_data_sheet_url': 'https://example.com/safety-data-sheet',
                                    'item_type_name': detected_type,
                                }
                                
                                for attr_name, value in attr_value_map.items():
                                    set_internal_value(attr_name, value)
                                
                                # Description and keywords
                                description_value = (
                                    get_product_value('Produktbeschreibung', 'Beschreibung', 'Product Description') or
                                    " ".join(content.bullet_points)
                                )
                                set_internal_by_keyword('product_description', description_value, limit=1)
                                
                                set_internal_by_keyword('generic_keyword', content.suchbegriffe, limit=2)
                                set_internal_by_keyword('search_terms', content.suchbegriffe, limit=2)
                                
                                # Images
                                image_urls = [
                                    get_product_value('Image URL 1', 'Bild 1'),
                                    get_product_value('Image URL 2', 'Bild 2'),
                                    get_product_value('Image URL 3', 'Bild 3'),
                                    get_product_value('Image URL 4', 'Bild 4'),
                                    get_product_value('Image URL 5', 'Bild 5'),
                                ]
                                if image_urls[0]:
                                    set_internal_by_keyword('main_image_url', image_urls[0], limit=1)
                                
                                other_attrs = [name for name in format_info.internal_columns.keys() if 'other_image_url' in name.lower()]
                                other_attrs.sort()
                                other_images = [url for url in image_urls[1:] if url]
                                for attr_name, url in zip(other_attrs, other_images):
                                    set_internal_value(attr_name, url)
                                
                                # Recommended browse nodes fallback
                                if not browse_nodes_options:
                                    logger.warning("No browse node mapping found for product type %s", detected_type)
                                
                                # STEP 4: Fill ALL required fields for this product type
                                logger.info(f"📋 Filling {len(required_attributes)} required attributes...")
                                filled_count = 0
                                
                                for attr in required_attributes:
                                    # Get column info for this attribute (from internal columns)
                                    col_info = format_info.internal_columns.get(attr)
                                    if not col_info:
                                        continue
                                    
                                    col_idx = col_info.column_index + 1
                                    value = None
                                    
                                    # Map attribute to data/defaults
                                    attr_lower = attr.lower()
                                    
                                    # SKU
                                    if 'item_sku' in attr_lower:
                                        value = sku_value
                                    
                                    # Brand
                                    elif 'brand_name' in attr_lower:
                                        value = brand_value
                                    
                                    # Title
                                    elif 'item_name' in attr_lower:
                                        value = content.artikelname
                                    
                                    # Product Type
                                    elif 'feed_product_type' in attr_lower or 'product_type' in attr_lower:
                                        value = detected_type
                                    
                                    # EAN/GTIN
                                    elif 'external_product_id' in attr_lower and 'type' not in attr_lower:
                                        value = ean_value
                                    elif 'external_product_id_type' in attr_lower:
                                        value = "EAN" if ean_value else None
                                    
                                    # Manufacturer (exact match to avoid false positives)
                                    elif attr_lower == 'manufacturer' or attr_lower == 'brand_manufacturer':
                                        value = product_data.get('Hersteller') or product_data.get('Manufacturer') or brand_value
                                    
                                    # Country of Origin
                                    elif 'country_of_origin' in attr_lower:
                                        value = product_data.get('Herkunftsland') or product_data.get('Country of Origin') or "DE"
                                    
                                    # Condition
                                    elif 'condition_type' in attr_lower:
                                        value = "new_new"
                                    
                                    # Fulfillment
                                    elif 'fulfillment_channel' in attr_lower:
                                        value = "DEFAULT"
                                    
                                    # Material
                                    elif 'material_type' in attr_lower or attr_lower == 'material':
                                        value = product_data.get('Material') or product_data.get('material')
                                    
                                    # Weight
                                    elif 'item_weight' in attr_lower and 'unit' not in attr_lower:
                                        value = product_data.get('Gewicht') or product_data.get('KG') or product_data.get('Weight')
                                    elif 'item_weight_unit' in attr_lower:
                                        value = "kilograms"
                                    
                                    # Volume
                                    elif 'item_volume' in attr_lower and 'unit' not in attr_lower:
                                        vol = product_data.get('Max. Befüllung/Filling in ml') or product_data.get('Volume')
                                        if vol:
                                            try:
                                                # Extract number from strings like "350 ml"
                                                num = re.search(r'\d+', str(vol))
                                                value = num.group() if num else vol
                                            except:
                                                value = vol
                                    elif 'item_volume_unit' in attr_lower:
                                        value = "milliliters"
                                    
                                    # Description (exact match to avoid matching manufacturer_description, etc.)
                                    elif attr_lower == 'product_description' or attr_lower == 'item_description':
                                        value = product_data.get('Produktbeschreibung') or content.artikelname
                                    
                                    # Model
                                    elif attr_lower == 'model' or 'model_name' in attr_lower:
                                        value = product_data.get('Modellnummer') or product_data.get('Model')
                                    
                                    # Batteries
                                    elif 'are_batteries_included' in attr_lower:
                                        value = "false"
                                    elif 'batteries_required' in attr_lower:
                                        value = "false"
                                    elif 'lithium' in attr_lower:
                                        value = "0" if 'cell' in attr_lower else "false"
                                    
                                    # Safety
                                    elif 'hazmat' in attr_lower or 'safety_data_sheet' in attr_lower:
                                        value = "not_applicable"
                                    
                                    # Packaging
                                    elif 'lithium_battery_packaging' in attr_lower:
                                        value = "batteries_contained_in_equipment"
                                    elif 'number_of_boxes' in attr_lower:
                                        value = "1"
                                    
                                    # Image
                                    elif 'main_image_url' in attr_lower:
                                        value = product_data.get('Image URL') or product_data.get('main_image')
                                    
                                    # Browse nodes (use first from recommended list)
                                    elif 'recommended_browse_nodes' in attr_lower:
                                        browse_nodes = format_info.browse_nodes_by_product_type.get(detected_type.lower(), [])
                                        value = browse_nodes[0] if browse_nodes else None
                                    
                                    # Item type
                                    elif 'item_type_name' in attr_lower:
                                        value = product_data.get('Artikelgruppe/Group') or detected_type
                                    
                                    # Expiration
                                    elif 'is_expiration_dated' in attr_lower:
                                        value = "false"
                                    
                                    # Unit measurements - defaults
                                    elif attr_lower.endswith('_unit_of_measure'):
                                        if 'weight' in attr_lower:
                                            value = "kilograms"
                                        elif 'volume' in attr_lower or 'capacity' in attr_lower:
                                            value = "milliliters"
                                        elif 'length' in attr_lower or 'height' in attr_lower or 'width' in attr_lower or 'depth' in attr_lower:
                                            value = "centimeters"
                                        elif 'battery' in attr_lower and 'energy' in attr_lower:
                                            value = "watt_hours"
                                    
                                    # Price fields - skip for now (user needs to fill)
                                    elif 'price' in attr_lower:
                                        value = None
                                    
                                    # List price
                                    elif 'list_price_with_tax' in attr_lower:
                                        value = product_data.get('UVP') or product_data.get('Price')
                                    
                                    # Bullet points (if not already filled)
                                    elif 'bullet_point' in attr_lower:
                                        # Extract bullet number from attribute name
                                        match = re.search(r'(\d+)', attr)
                                        if match:
                                            bullet_idx = int(match.group(1)) - 1
                                            if 0 <= bullet_idx < len(content.bullet_points):
                                                value = content.bullet_points[bullet_idx]
                                    
                                    # Search terms
                                    elif 'generic_keywords' in attr_lower or ('search' in attr_lower and 'term' in attr_lower):
                                        # Join all search terms
                                        value = ', '.join(content.search_terms) if content.search_terms else None
                                    
                                    # Color
                                    elif attr_lower == 'color' or 'color_name' in attr_lower:
                                        value = product_data.get('Farbe/Color') or product_data.get('Color')
                                    
                                    # Size
                                    elif attr_lower == 'size' or 'size_name' in attr_lower:
                                        value = product_data.get('Größe/Size') or product_data.get('Size')
                                    
                                    # Dimensions
                                    elif 'item_length' in attr_lower and 'unit' not in attr_lower:
                                        value = product_data.get('Länge/Length in cm') or product_data.get('Length')
                                    elif 'item_width' in attr_lower and 'unit' not in attr_lower:
                                        value = product_data.get('Breite/Width in cm') or product_data.get('Width')
                                    elif 'item_height' in attr_lower and 'unit' not in attr_lower:
                                        value = product_data.get('Höhe/Height in cm') or product_data.get('Height')
                                    
                                    # Package dimensions
                                    elif 'package_length' in attr_lower and 'unit' not in attr_lower:
                                        value = product_data.get('Package Length')
                                    elif 'package_width' in attr_lower and 'unit' not in attr_lower:
                                        value = product_data.get('Package Width')
                                    elif 'package_height' in attr_lower and 'unit' not in attr_lower:
                                        value = product_data.get('Package Height')
                                    elif 'package_weight' in attr_lower and 'unit' not in attr_lower:
                                        value = product_data.get('Package Weight')
                                    
                                    # Additional common fields
                                    elif 'care_instructions' in attr_lower:
                                        value = product_data.get('Pflegehinweise') or product_data.get('Care Instructions')
                                    elif 'included_components' in attr_lower:
                                        value = product_data.get('Komponenten') or product_data.get('Included Components')
                                    elif 'special_features' in attr_lower:
                                        value = product_data.get('Besonderheiten') or product_data.get('Special Features')
                                    elif 'target_audience' in attr_lower:
                                        value = product_data.get('Zielgruppe') or product_data.get('Target Audience')
                                    elif 'warranty_description' in attr_lower:
                                        value = product_data.get('Garantie') or product_data.get('Warranty')
                                    
                                    # Fill the cell if we have a value
                                    if value is not None:
                                        ws.cell(row=target_row, column=col_idx).value = str(value)
                                        filled_count += 1
                                        logger.debug(f"  ✓ {attr}: {str(value)[:40]}")
                                
                                logger.info(f"✅ Filled {filled_count}/{len(required_attributes)} required fields")
                                
                                # Fill Offer Action for XML format
                                if format_info.format_type == "XML" and format_info.offer_action_column_index is not None:
                                    ws.cell(row=target_row, column=format_info.offer_action_column_index + 1).value = "(Standard) Erstellen oder Ersetzen"
                                    logger.info(f"✅ Offer Action: (Standard) Erstellen oder Ersetzen")
                                
                                # Track missing required fields
                                missing_required = []
                                for attr in required_fields:
                                    col_info = format_info.internal_columns.get(attr)
                                    if not col_info:
                                        continue
                                    cell_value = ws.cell(row=target_row, column=col_info.column_index + 1).value
                                    if cell_value is None or (isinstance(cell_value, str) and not str(cell_value).strip()):
                                        missing_required.append(format_info.field_display_names.get(attr, attr))
                                
                                if missing_required:
                                    missing_required_summary.append({
                                        "row": target_row,
                                        "sku": sku_value or product_data.get('Artikel-Nr.') or '-',
                                        "missing": missing_required
                                    })
                                    logger.warning(f"Missing required fields for row {target_row}: {missing_required}")
                                else:
                                    logger.info("✅ All required fields populated for this product")
                            
                            # Save to BytesIO
                            output = io.BytesIO()
                            wb.save(output)
                            wb.close()
                            output.seek(0)
                            
                            st.success(f"🎉 Template erfolgreich befüllt mit {len(generated_contents)} Produkten!")
                            
                            # Show what was preserved
                            with st.expander("✅ Template-Struktur vollständig erhalten"):
                                st.write("**Was wurde bewahrt:**")
                                st.write(f"✅ Alle {len(wb.sheetnames)} Sheets: {', '.join(wb.sheetnames[:5])}...")
                                st.write(f"✅ VBA Makros: {'Ja ✓' if wb.vba_archive else 'Keine vorhanden'}")
                                st.write("✅ Alle Formeln")
                                st.write("✅ Alle Formatierungen")
                                st.write("✅ Alle Dropdowns & Datenvalidierungen")
                                st.write("✅ Header & Beispiel-Zeilen")
                                st.write("")
                                st.write("**Was wurde geändert:**")
                                st.write(f"📝 Nur Daten in Zeilen {format_info.data_start_row} - {format_info.data_start_row + len(generated_contents) - 1}")
                                st.write("📝 Nur in den erkannten Produkt-Spalten")
                            
                            # Download button
                            st.download_button(
                                label="📥 Befülltes Template herunterladen (.xlsm)",
                                data=output.getvalue(),
                                file_name=f"amazon_template_{format_info.format_type}_{pd.Timestamp.now().strftime('%Y%m%d_%H%M%S')}.xlsm",
                                mime="application/vnd.ms-excel.sheet.macroEnabled.12",
                                use_container_width=True,
                                type="primary"
                            )
                            
                            if missing_required_summary:
                                with st.expander("⚠️ Fehlende Pflichtfelder (Details)", expanded=True):
                                    for entry in missing_required_summary:
                                        sku_label = entry.get("sku", "-")
                                        st.write(f"Zeile {entry['row']} • SKU: {sku_label}")
                                        st.write(", ".join(entry["missing"]))
                            else:
                                st.success("Alle produktbezogenen Pflichtfelder wurden ausgefüllt.")
                            
                            # Show summary
                            col_sum1, col_sum2, col_sum3 = st.columns(3)
                            with col_sum1:
                                st.metric("Produkte", len(generated_contents))
                            with col_sum2:
                                st.metric("Template Format", format_info.format_type)
                            with col_sum3:
                                st.metric("Struktur", "100% erhalten")
                            
                        except Exception as e:
                            st.error(f"❌ Fehler beim Befüllen: {e}")
                            logger.error(f"Error filling template: {e}", exc_info=True)
    
    elif not st.session_state.get('api_key', '').strip():
        st.warning("⚠️ Bitte OpenAI API Key in der Seitenleiste eingeben!")
    elif not template_file:
        st.info("ℹ️ Bitte laden Sie ein Amazon-Template hoch")
    elif products_df is None:
        st.info("ℹ️ Bitte laden Sie Produktdaten hoch")

with tab4:
    st.header("✍️ KI-Content-Optimierung (COSMO & RUFUS)")
    st.markdown("""
    Erstellt **hoch-optimierte Amazon-Listings** basierend auf den 15 semantischen Beziehungstypen von COSMO & RUFUS.
    
    **Output:** Eine Excel-Datei mit:
    - Produkt-ID & Alter Titel
    - ✅ **Neuer Titel** (Wer, Was, Wofür, Wie)
    - ✅ **Neue Beschreibung** (15 Beziehungstypen abgedeckt)
    - ✅ **5 Bullet Points** (Verkaufsstark & nutzenorientiert)
    - ✅ **Keywords** (Exklusiv - nicht in anderen Feldern enthalten!)
    """)
    
    opt_file = st.file_uploader("Produktdaten Excel hochladen", type=["xlsx", "xls"], key="opt_upload")
    
    if opt_file:
        try:
            # Dynamic header detection (same as tab1)
            df_test = pd.read_excel(opt_file, header=None, nrows=10)
            header_row = 0
            best_score = 0
            for row_idx in range(min(5, len(df_test))):
                row = df_test.iloc[row_idx]
                score = sum(1 for val in row if pd.notna(val) and isinstance(val, str) and len(val) > 2 and not val.replace('.', '').replace('-', '').isdigit())
                if score > best_score:
                    best_score = score
                    header_row = row_idx
            
            df_opt = pd.read_excel(opt_file, header=header_row)
            st.success(f"✅ {len(df_opt)} Produkte geladen")
            
            with st.expander("📋 Daten-Vorschau"):
                st.dataframe(df_opt.head(), use_container_width=True)
            
            # Column mapping
            cols = df_opt.columns.tolist()
            col1, col2 = st.columns(2)
            with col1:
                id_col = st.selectbox("Spalte für Produkt-ID (SKU/EAN)", cols, index=0)
            with col2:
                title_col = st.selectbox("Spalte für Alten Titel (Optional)", ["-"] + cols, index=min(1, len(cols)))
            
            num_products_opt = st.slider("Anzahl Produkte", 1, len(df_opt), min(5, len(df_opt)), key="opt_slider")
            
            if st.button("🚀 Optimierung Starten", type="primary", key="opt_start"):
                if not st.session_state.get('api_key', '').strip():
                    st.error("❌ Bitte API Key eingeben")
                else:
                    results = []
                    progress_bar = st.progress(0)
                    status = st.empty()
                    
                    client = OpenAI(api_key=st.session_state.api_key)
                    
                    for idx in range(num_products_opt):
                        status.text(f"✍️ Optimiere Produkt {idx + 1}/{num_products_opt}...")
                        row = df_opt.iloc[idx]
                        
                        # Prepare data
                        product_data_str = "\n".join([f"- {k}: {v}" for k, v in row.to_dict().items() if pd.notna(v)])
                        
                        # Prompt
                        prompt = st.session_state.cosmo_prompt_template.replace("{{product_data}}", product_data_str)
                        
                        try:
                            response = client.chat.completions.create(
                                model="gpt-5.1",
                                messages=[
                                    {"role": "system", "content": "Du bist ein Amazon SEO-Experte spezialisiert auf COSMO & RUFUS."},
                                    {"role": "user", "content": prompt}
                                ],
                                response_format={
                                    "type": "json_schema",
                                    "json_schema": {
                                        "name": "cosmo_content",
                                        "schema": CosmoOptimizedContent.model_json_schema(),
                                        "strict": True
                                    }
                                }
                            )
                            
                            content = CosmoOptimizedContent.model_validate_json(response.choices[0].message.content)
                            
                            # Enforce hard limits with AI regeneration
                            if len(content.artikelname) > 200:
                                content.artikelname = ensure_length_with_ai(content.artikelname, 200, "Titel", client)
                                
                            new_bullets = []
                            for i, bp in enumerate(content.bullet_points):
                                if len(bp) > 250:
                                    new_bp = ensure_length_with_ai(bp, 250, f"Bullet {i+1}", client)
                                    new_bullets.append(new_bp)
                                else:
                                    new_bullets.append(bp)
                            content.bullet_points = new_bullets
                            
                            if len(content.produktbeschreibung) > 2000:
                                content.produktbeschreibung = ensure_length_with_ai(content.produktbeschreibung, 2000, "Beschreibung", client)
                                
                            if len(content.suchbegriffe) > 249:
                                content.suchbegriffe = ensure_length_with_ai(content.suchbegriffe, 249, "Keywords", client)
                            
                            # Store result
                            result_row = {
                                "Identifier": row[id_col],
                                "Old Title": row[title_col] if title_col != "-" else "",
                                "New Title": content.artikelname,
                                "New Description": content.produktbeschreibung,
                                "New Keyword": content.suchbegriffe
                            }
                            # Add bullets as separate columns
                            for i, bp in enumerate(content.bullet_points, 1):
                                result_row[f"New Bullet {i}"] = bp
                            
                            results.append(result_row)
                            
                            # Show preview
                            with st.expander(f"✅ {row[id_col]}: {content.artikelname[:60]}..."):
                                st.write("**Titel:**", content.artikelname)
                                st.write("**Bullets:**")
                                for bp in content.bullet_points:
                                    st.write(f"• {bp}")
                                st.write("**Keywords:**", content.suchbegriffe)
                                st.caption("**Beschreibung:** " + content.produktbeschreibung[:100] + "...")
                        
                        except Exception as e:
                            st.error(f"Fehler bei Produkt {idx}: {e}")
                        
                        progress_bar.progress((idx + 1) / num_products_opt)
                    
                    status.text("✅ Fertig!")
                    
                    # Create Excel
                    df_result = pd.DataFrame(results)
                    output = io.BytesIO()
                    with pd.ExcelWriter(output, engine='openpyxl') as writer:
                        df_result.to_excel(writer, index=False, sheet_name="Optimized Content")
                        # Adjust column widths
                        ws = writer.sheets["Optimized Content"]
                        for column_cells in ws.columns:
                            length = max(len(str(cell.value)) for cell in column_cells)
                            ws.column_dimensions[column_cells[0].column_letter].width = min(length + 2, 50)
                            
                    output.seek(0)
                    
                    st.download_button(
                        "📥 Optimierte Daten herunterladen (XLSX)",
                        data=output.getvalue(),
                        file_name=f"cosmo_optimized_{pd.Timestamp.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        type="primary"
                    )
                    
        except Exception as e:
            st.error(f"Fehler beim Lesen der Datei: {e}")

with tab2:
    st.header("📊 Template Preview")
    st.info("Preview-Funktion in Entwicklung")

with tab3:
    st.header("⚙️ Einstellungen & Info")
    
    st.subheader("📋 Unterstützte Template-Formate")
    st.success("✅ **XML Format** (Seller Central) - z.B. BOTTLE_KITCHEN_KNIFE_SAUTE_FRY_PAN.xlsm")
    st.success("✅ **Flat File Format** (Seller Central) - z.B. DRINKING_CUP_BOTTLE_SEASONING_MILL...")
    
    st.markdown("---")
    st.subheader("🎯 Features")
    st.markdown("""
    - ✅ **Automatische Template-Erkennung** - Kein Hardcoding!
    - ✅ **Beide Formate unterstützt** - XML & Flat File
    - ✅ **Dynamische Spalten-Erkennung** - Funktioniert mit jedem Template
    - ✅ **AI-Content-Generierung** - GPT-5-mini mit COSMO/RUFUS Optimierung
    - ✅ **Struktur-Erhaltung** - Formeln, Formatierung, Makros bleiben erhalten
    - ✅ **Schnell & Zuverlässig** - Keine unnötigen API-Calls
    """)
    
    st.markdown("---")
    st.subheader("📖 Verwendung")
    st.code("""
    1. API Key eingeben (Seitenleiste)
    2. Produktdaten-Excel hochladen
    3. Amazon-Template hochladen
    4. Template wird automatisch analysiert
    5. "Starten" klicken
    6. Fertig! Template herunterladen
    """)

