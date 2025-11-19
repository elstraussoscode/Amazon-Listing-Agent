"""
Dynamic Amazon Template Analyzer
=================================

Analyzes ANY Amazon template file to detect:
- Template format (XML vs Flat File)
- Header row location
- Data start row
- Column mappings
- Product type options
- Required fields per product type

Works with ANY Amazon Seller/Vendor template!
"""

import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
from typing import Dict, List, Optional, Tuple, Set
from dataclasses import dataclass, field
from pydantic import BaseModel, Field as PydanticField
import logging

logger = logging.getLogger(__name__)


class ColumnInfo(BaseModel):
    """Information about a detected column"""
    column: str = PydanticField(description="Column letter (e.g., 'D', 'CC')")
    column_index: int = PydanticField(description="Zero-based column index")
    row: Optional[int] = PydanticField(default=None, description="Row number where header is found")


@dataclass
class TemplateFormat:
    """Detected template format information"""
    format_type: str  # "XML" or "FLAT_FILE"
    sheet_name: str  # "Vorlage" or other
    header_row: int
    internal_row: Optional[int]
    example_row: Optional[int]
    data_start_row: int
    
    # Key columns
    sku_column: str
    sku_column_index: int
    product_type_column: str
    product_type_column_index: int
    brand_column: Optional[str]
    brand_column_index: Optional[int]
    title_column: Optional[str]
    title_column_index: Optional[int]
    
    # For XML format only
    offer_action_column: Optional[str] = None
    offer_action_column_index: Optional[int] = None
    
    # For Flat File format only
    update_delete_column: Optional[str] = None
    update_delete_column_index: Optional[int] = None
    
    # Product types available
    product_types: List[str] = None
    
    # Total columns
    total_columns: int = 0
    
    # All detected columns with display headers
    key_columns: Dict[str, ColumnInfo] = field(default_factory=dict)
    
    # Internal attribute name -> column info (from internal header row)
    internal_columns: Dict[str, ColumnInfo] = field(default_factory=dict)
    
    # Required vs Optional fields from Datendefinitionen (display names)
    required_fields: List[str] = field(default_factory=list)
    optional_fields: List[str] = field(default_factory=list)
    
    # Internal attribute names that are globally required
    required_internal_fields: Set[str] = field(default_factory=set)
    
    # Mapping internal attribute -> human readable label
    field_display_names: Dict[str, str] = field(default_factory=dict)
    
    # Product-type specific required internal attributes
    product_type_required_attributes: Dict[str, List[str]] = field(default_factory=dict)
    
    # Browse node options per product type
    browse_nodes_by_product_type: Dict[str, List[str]] = field(default_factory=dict)


class DynamicTemplateAnalyzer:
    """Analyzes Amazon templates dynamically"""
    
    def __init__(self, template_path: str):
        self.template_path = template_path
        self.wb = None
        self.format_info: Optional[TemplateFormat] = None
    
    def analyze(self) -> TemplateFormat:
        """Main analysis method"""
        logger.info(f"Analyzing template: {self.template_path}")
        
        self.wb = openpyxl.load_workbook(self.template_path, data_only=False, keep_vba=True)
        
        # Step 1: Find the data entry sheet
        sheet_name = self._find_data_sheet()
        logger.info(f"Found data sheet: {sheet_name}")
        
        ws = self.wb[sheet_name]
        
        # Step 2: Detect template format and structure
        format_info = self._detect_format(ws, sheet_name)
        
        # Step 3: Find product types
        product_types = self._find_product_types()
        format_info.product_types = product_types
        
        # Step 4: Parse required/optional fields from Datendefinitionen
        (
            required_fields,
            optional_fields,
            required_internal_fields,
            field_display_map,
        ) = self._parse_data_definitions()
        format_info.required_fields = required_fields
        format_info.optional_fields = optional_fields
        format_info.required_internal_fields = required_internal_fields
        format_info.field_display_names = field_display_map
        
        # Step 5: Parse attribute/product type map and browse nodes
        product_type_requirements = self._parse_attribute_ptd_map(required_internal_fields)
        format_info.product_type_required_attributes = product_type_requirements
        format_info.browse_nodes_by_product_type = self._parse_browse_nodes()
        
        self.format_info = format_info
        
        logger.info(f"Template format detected: {format_info.format_type}")
        logger.info(f"Header row: {format_info.header_row}")
        logger.info(f"Data starts: {format_info.data_start_row}")
        logger.info(f"Product types found: {len(product_types)}")
        logger.info(f"Required fields: {len(required_fields)}")
        logger.info(f"Optional fields: {len(optional_fields)}")
        
        return format_info
    
    def _find_data_sheet(self) -> str:
        """Find the main data entry sheet (Vorlage)"""
        
        # Check for "Vorlage" sheet
        if 'Vorlage' in self.wb.sheetnames:
            return 'Vorlage'
        
        # Check for sheets with "Vorlage" in name
        for sheet_name in self.wb.sheetnames:
            if 'vorlage' in sheet_name.lower():
                return sheet_name
        
        # Fallback: use first sheet with significant data
        for sheet_name in self.wb.sheetnames:
            ws = self.wb[sheet_name]
            if ws.max_row > 10 and ws.max_column > 10:
                return sheet_name
        
        raise ValueError("Could not find data entry sheet in template")
    
    def _detect_format(self, ws, sheet_name: str) -> TemplateFormat:
        """Detect template format by analyzing structure"""
        
        # Strategy: Check multiple rows to find headers
        # XML format typically has headers in Row 4
        # Flat File format has headers in Row 2
        
        for check_row in [2, 4, 1, 3, 5]:
            row_data = {}
            for col_idx in range(1, ws.max_column + 1):
                val = ws.cell(row=check_row, column=col_idx).value
                if val is not None and str(val).strip():
                    row_data[col_idx] = str(val).strip()
            
            if len(row_data) > 5:  # Likely a header row if many columns have data
                return self._analyze_header_row(ws, sheet_name, check_row, row_data)
        
        raise ValueError("Could not detect template format - no clear header row found")
    
    def _analyze_header_row(self, ws, sheet_name: str, header_row: int, row_data: Dict[int, str]) -> TemplateFormat:
        """Analyze a potential header row to extract column mappings"""
        
        logger.info(f"Analyzing header row: {header_row}")
        logger.info(f"Columns found: {list(row_data.values())[:10]}")
        
        # Detect format by looking for key columns
        format_type = None
        sku_col = None
        product_type_col = None
        brand_col = None
        title_col = None
        offer_action_col = None
        update_delete_col = None
        
        # Search for key columns
        for col_idx, header_text in row_data.items():
            header_lower = header_text.lower()
            
            # SKU - prioritize column A or B, exact matches
            if 'sku' in header_lower and 'vendor' not in header_lower:
                # Prioritize exact matches and early columns
                if not sku_col or header_lower in ['sku', 'seller-sku', 'verkäufer-sku'] or col_idx <= 2:
                    sku_col = col_idx
            
            # Product Type
            if 'produkttyp' in header_lower or ('product' in header_lower and 'type' in header_lower):
                product_type_col = col_idx
            
            # Brand
            if 'marke' in header_lower or 'brand' in header_lower:
                brand_col = col_idx
            
            # Title
            if 'artikelname' in header_lower or ('item' in header_lower and 'name' in header_lower) or 'produktname' in header_lower:
                title_col = col_idx
            
            # Offer Action (XML format)
            if 'angebotsaktion' in header_lower or 'offer' in header_lower and 'action' in header_lower:
                offer_action_col = col_idx
                format_type = "XML"
            
            # Update/Delete (Flat File format)
            if ('update' in header_lower and 'delete' in header_lower) or 'update_delete' in header_lower:
                update_delete_col = col_idx
                format_type = "FLAT_FILE"
        
        # Build key_columns dictionary with all detected columns
        key_columns = {}
        
        for col_idx, header_text in row_data.items():
            header_lower = header_text.lower()
            
            # EAN/GTIN/Barcode
            if ('ean' in header_lower or 'gtin' in header_lower or 'barcode' in header_lower) and 'EAN' not in key_columns:
                key_columns['EAN'] = ColumnInfo(column=get_column_letter(col_idx), column_index=col_idx - 1, row=header_row)
            
            # Material
            if 'material' in header_lower and 'Material' not in key_columns:
                key_columns['Material'] = ColumnInfo(column=get_column_letter(col_idx), column_index=col_idx - 1, row=header_row)
            
            # Color/Colour/Farbe
            if ('farbe' in header_lower or 'color' in header_lower or 'colour' in header_lower) and 'Color' not in key_columns:
                key_columns['Color'] = ColumnInfo(column=get_column_letter(col_idx), column_index=col_idx - 1, row=header_row)
            
            # Size/Größe/Dimensions
            if ('größe' in header_lower or 'size' in header_lower or 'dimension' in header_lower) and 'Size' not in key_columns:
                key_columns['Size'] = ColumnInfo(column=get_column_letter(col_idx), column_index=col_idx - 1, row=header_row)
            
            # Weight/Gewicht
            if ('gewicht' in header_lower or 'weight' in header_lower) and 'Weight' not in key_columns:
                key_columns['Weight'] = ColumnInfo(column=get_column_letter(col_idx), column_index=col_idx - 1, row=header_row)
            
            # Description/Beschreibung
            if ('beschreibung' in header_lower or 'description' in header_lower) and 'Description' not in key_columns:
                key_columns['Description'] = ColumnInfo(column=get_column_letter(col_idx), column_index=col_idx - 1, row=header_row)
            
            # Manufacturer/Hersteller
            if ('hersteller' in header_lower or 'manufacturer' in header_lower) and 'Manufacturer' not in key_columns:
                key_columns['Manufacturer'] = ColumnInfo(column=get_column_letter(col_idx), column_index=col_idx - 1, row=header_row)
            
            # Bullet Points (find all - numbered)
            if 'aufzählungspunkt' in header_lower or 'bullet' in header_lower:
                import re
                num_match = re.search(r'\d+', header_text)
                if num_match:
                    bp_num = int(num_match.group())
                    key_columns[f'Bullet Point {bp_num}'] = ColumnInfo(column=get_column_letter(col_idx), column_index=col_idx - 1, row=header_row)
            
            # Search Terms (find all - numbered)
            if 'suchbegriff' in header_lower or 'search term' in header_lower or 'generic_keyword' in header_lower:
                import re
                num_match = re.search(r'\d+', header_text)
                if num_match:
                    st_num = int(num_match.group())
                    key_columns[f'Search Term {st_num}'] = ColumnInfo(column=get_column_letter(col_idx), column_index=col_idx - 1, row=header_row)
        
        # Add basic columns to key_columns
        if sku_col:
            key_columns['SKU'] = ColumnInfo(column=get_column_letter(sku_col), column_index=sku_col - 1, row=header_row)
        if product_type_col:
            key_columns['Product Type'] = ColumnInfo(column=get_column_letter(product_type_col), column_index=product_type_col - 1, row=header_row)
        if brand_col:
            key_columns['Brand'] = ColumnInfo(column=get_column_letter(brand_col), column_index=brand_col - 1, row=header_row)
        if title_col:
            key_columns['Title'] = ColumnInfo(column=get_column_letter(title_col), column_index=title_col - 1, row=header_row)
        if offer_action_col:
            key_columns['Offer Action'] = ColumnInfo(column=get_column_letter(offer_action_col), column_index=offer_action_col - 1, row=header_row)
        if update_delete_col:
            key_columns['Update/Delete'] = ColumnInfo(column=get_column_letter(update_delete_col), column_index=update_delete_col - 1, row=header_row)
        
        # Determine format if not yet determined
        if not format_type:
            # XML format typically has: SKU in col A or B, Product Type in col C
            # Flat File: Product Type in col A, SKU in col B
            if product_type_col and product_type_col == 1:  # Column A
                format_type = "FLAT_FILE"
            else:
                format_type = "XML"
        
        # Determine data start row and other structural elements
        if format_type == "XML":
            # XML: header row 4, internal row 5, example row 6, data starts row 7
            internal_row = header_row + 1 if header_row == 4 else None
            example_row = header_row + 2 if header_row == 4 else None
            data_start_row = header_row + 3 if header_row == 4 else header_row + 1
        else:
            # Flat File: header row 2, internal row 3, data starts row 4
            internal_row = header_row + 1 if header_row == 2 else None
            example_row = None
            data_start_row = header_row + 2 if header_row == 2 else header_row + 1
        
        # Map internal attribute names to columns (next row after headers)
        internal_columns = {}
        internal_row_idx = internal_row or (header_row + 1)
        if internal_row_idx <= ws.max_row:
            for col_idx in row_data.keys():
                internal_name = ws.cell(row=internal_row_idx, column=col_idx).value
                if internal_name:
                    key = str(internal_name).strip()
                    internal_columns[key] = ColumnInfo(
                        column=get_column_letter(col_idx),
                        column_index=col_idx - 1,
                        row=internal_row_idx,
                    )
        
        # Build TemplateFormat object
        return TemplateFormat(
            format_type=format_type,
            sheet_name=sheet_name,
            header_row=header_row,
            internal_row=internal_row,
            example_row=example_row,
            data_start_row=data_start_row,
            sku_column=get_column_letter(sku_col) if sku_col else None,
            sku_column_index=sku_col - 1 if sku_col else None,
            product_type_column=get_column_letter(product_type_col) if product_type_col else None,
            product_type_column_index=product_type_col - 1 if product_type_col else None,
            brand_column=get_column_letter(brand_col) if brand_col else None,
            brand_column_index=brand_col - 1 if brand_col else None,
            title_column=get_column_letter(title_col) if title_col else None,
            title_column_index=title_col - 1 if title_col else None,
            offer_action_column=get_column_letter(offer_action_col) if offer_action_col else None,
            offer_action_column_index=offer_action_col - 1 if offer_action_col else None,
            update_delete_column=get_column_letter(update_delete_col) if update_delete_col else None,
            update_delete_column_index=update_delete_col - 1 if update_delete_col else None,
            total_columns=ws.max_column,
            key_columns=key_columns,
            internal_columns=internal_columns,
        )
    
    def _find_product_types(self) -> List[str]:
        """Find all available product types from named ranges or dropdowns"""
        
        product_types = []
        
        # Look for named range with product types
        # XML templates: "product_type1.value"
        # Flat File templates: "feed_product_type"
        for name in self.wb.defined_names:
            name_lower = name.lower()
            is_product_type_range = (
                ('product' in name_lower and 'type' in name_lower and 'value' in name_lower) or
                (name_lower == 'feed_product_type')
            )
            
            if is_product_type_range:
                # Found product type range
                range_ref = self.wb.defined_names[name].value
                
                if not range_ref or '!' not in range_ref:
                    continue
                
                # Parse range reference
                sheet_name = range_ref.split('!')[0].strip("'")
                range_part = range_ref.split('!')[-1]
                
                if sheet_name not in self.wb.sheetnames:
                    continue
                
                ws = self.wb[sheet_name]
                
                # Parse column and row
                if '$' in range_part:
                    parts = range_part.split('$')
                    col_letter = parts[1]
                    row_start = int(parts[2].split(':')[0])
                    
                    col_idx = column_index_from_string(col_letter)
                    
                    # Read values
                    for row_idx in range(row_start, min(row_start + 200, ws.max_row + 1)):
                        val = ws.cell(row=row_idx, column=col_idx).value
                        if val and str(val).strip():
                            pt = str(val).strip()
                            if pt not in product_types:
                                product_types.append(pt)
                        elif len(product_types) > 3:
                            # Check for consecutive empty cells
                            empty_count = sum(
                                1 for r in range(row_idx, min(row_idx + 3, ws.max_row + 1))
                                if not ws.cell(row=r, column=col_idx).value
                            )
                            if empty_count >= 3:
                                break
                
                if product_types:
                    break
        
        return product_types
    
    def _parse_data_definitions(self) -> Tuple[List[str], List[str], Set[str], Dict[str, str]]:
        """Parse Datendefinitionen sheet to extract required/optional fields"""
        
        required_fields: List[str] = []
        optional_fields: List[str] = []
        required_internal_fields: Set[str] = set()
        field_display_map: Dict[str, str] = {}
        
        try:
            if 'Datendefinitionen' not in self.wb.sheetnames:
                logger.warning("Datendefinitionen sheet not found")
                return required_fields, optional_fields, required_internal_fields, field_display_map
            
            ws = self.wb['Datendefinitionen']
            header_row = 2  # Actual headers are on row 2
            max_row = ws.max_row
            max_col = ws.max_column
            
            def find_column(keyword: str) -> Optional[int]:
                keyword_lower = keyword.lower()
                for col in range(1, max_col + 1):
                    val = ws.cell(row=header_row, column=col).value
                    if val and keyword_lower in str(val).lower():
                        return col
                return None
            
            pflicht_col = find_column('pflichtfeld')
            local_col = find_column('lokale')
            fieldname_col = find_column('feldname')
            
            if not pflicht_col or not local_col or not fieldname_col:
                logger.warning("Unable to locate Pflichtfeld/Lokale Bezeichnung/Feldname columns")
                return required_fields, optional_fields, required_internal_fields, field_display_map
            
            for row in range(header_row + 1, max_row + 1):
                pf_value = ws.cell(row=row, column=pflicht_col).value
                local_value = ws.cell(row=row, column=local_col).value
                field_value = ws.cell(row=row, column=fieldname_col).value
                
                if not pf_value and not local_value:
                    continue
                
                pf_text = str(pf_value).strip().lower() if pf_value else ''
                local_text = str(local_value).strip() if local_value else None
                field_name = str(field_value).strip() if field_value else None
                
                if field_name and local_text:
                    field_display_map[field_name] = local_text
                
                if not pf_text:
                    continue
                
                if any(token in pf_text for token in ['pflicht', 'erforder']):
                    if local_text:
                        required_fields.append(local_text)
                    if field_name:
                        required_internal_fields.add(field_name)
                elif 'optional' in pf_text:
                    if local_text:
                        optional_fields.append(local_text)
            
            logger.info(
                f"Parsed {len(required_fields)} required fields, "
                f"{len(optional_fields)} optional fields, "
                f"{len(required_internal_fields)} required internal attributes"
            )
            
        except Exception as e:
            logger.error(f"Error parsing Datendefinitionen: {e}")
        
        return required_fields, optional_fields, required_internal_fields, field_display_map
    
    def _parse_attribute_ptd_map(self, required_internal_fields: Set[str]) -> Dict[str, List[str]]:
        """Parse AttributePTDMAP sheet to get product type specific required attributes"""
        mapping: Dict[str, List[str]] = {}
        
        try:
            if 'AttributePTDMAP' not in self.wb.sheetnames:
                return mapping
            
            ws = self.wb['AttributePTDMAP']
            max_row = ws.max_row
            max_col = ws.max_column
            
            product_types: List[Tuple[int, str]] = []
            for col in range(2, max_col + 1):
                pt_value = ws.cell(row=1, column=col).value
                if pt_value:
                    product_types.append((col, str(pt_value).strip().lower()))
            
            if not product_types:
                return mapping
            
            for row in range(2, max_row + 1):
                attr_value = ws.cell(row=row, column=1).value
                if not attr_value:
                    continue
                attribute = str(attr_value).strip()
                
                if attribute not in required_internal_fields:
                    continue
                
                for col_idx, pt in product_types:
                    cell_value = ws.cell(row=row, column=col_idx).value
                    if cell_value is None:
                        continue
                    
                    include = False
                    if isinstance(cell_value, (int, float)):
                        include = cell_value != 0
                    else:
                        include = str(cell_value).strip().lower() not in ('', '0', 'no', 'false', 'nein')
                    
                    if include:
                        mapping.setdefault(pt, []).append(attribute)
            
        except Exception as e:
            logger.error(f"Error parsing AttributePTDMAP: {e}")
        
        return mapping
    
    def _parse_browse_nodes(self) -> Dict[str, List[str]]:
        """Parse Gültige Werte sheet to map browse nodes per product type"""
        browse_nodes: Dict[str, List[str]] = {}
        
        try:
            if 'Gültige Werte' not in self.wb.sheetnames:
                return browse_nodes
            
            import re
            
            ws = self.wb['Gültige Werte']
            for row in ws.iter_rows(values_only=True):
                if not row or len(row) < 3:
                    continue
                label = row[1]
                if not label:
                    continue
                label_text = str(label)
                if 'produktkategorisierung' not in label_text.lower():
                    continue
                
                match = re.search(r'\[([^\]]+)\]', label_text)
                if not match:
                    continue
                
                product_type = match.group(1).strip().lower()
                values = [str(val).strip() for val in row[2:] if val and str(val).strip()]
                if values:
                    browse_nodes[product_type] = values
        
        except Exception as e:
            logger.error(f"Error parsing browse nodes: {e}")
        
        return browse_nodes
    
    def get_example_product_type(self) -> Optional[str]:
        """Get the example product type from the template"""
        if not self.format_info or not self.format_info.example_row:
            return None
        
        ws = self.wb[self.format_info.sheet_name]
        col_idx = self.format_info.product_type_column_index + 1
        val = ws.cell(row=self.format_info.example_row, column=col_idx).value
        
        return str(val).strip() if val else None
    
    def get_required_fields_for_product_type(self, product_type: str) -> List[str]:
        """Get required field display names for a specific product type"""
        if not product_type or not self.format_info:
            return []
        
        product_type_lower = product_type.strip().lower()
        required_attributes = self.format_info.product_type_required_attributes.get(product_type_lower, [])
        display_names = []
        for attr in required_attributes:
            display = self.format_info.field_display_names.get(attr, attr)
            display_names.append(display)
        return display_names
    
    def close(self):
        """Close the workbook"""
        if self.wb:
            self.wb.close()


def analyze_template(template_path: str) -> TemplateFormat:
    """
    Quick function to analyze a template file
    
    Usage:
        format_info = analyze_template("path/to/template.xlsm")
        print(f"Format: {format_info.format_type}")
        print(f"SKU column: {format_info.sku_column}")
        print(f"Product types: {format_info.product_types}")
    """
    analyzer = DynamicTemplateAnalyzer(template_path)
    format_info = analyzer.analyze()
    analyzer.close()
    return format_info


# Example usage
if __name__ == "__main__":
    import sys
    
    logging.basicConfig(level=logging.INFO)
    
    if len(sys.argv) > 1:
        template_path = sys.argv[1]
        format_info = analyze_template(template_path)
        
        print("\n" + "="*80)
        print("TEMPLATE ANALYSIS RESULT")
        print("="*80)
        print(f"Format Type: {format_info.format_type}")
        print(f"Sheet Name: {format_info.sheet_name}")
        print(f"Header Row: {format_info.header_row}")
        print(f"Data Starts: {format_info.data_start_row}")
        print(f"\nKey Columns:")
        print(f"  SKU: {format_info.sku_column} (index {format_info.sku_column_index})")
        print(f"  Product Type: {format_info.product_type_column} (index {format_info.product_type_column_index})")
        print(f"  Brand: {format_info.brand_column}")
        print(f"  Title: {format_info.title_column}")
        
        if format_info.format_type == "XML":
            print(f"  Offer Action: {format_info.offer_action_column}")
        else:
            print(f"  Update/Delete: {format_info.update_delete_column}")
        
        print(f"\nProduct Types ({len(format_info.product_types)}):")
        for i, pt in enumerate(format_info.product_types[:20], 1):
            print(f"  {i}. {pt}")
        if len(format_info.product_types) > 20:
            print(f"  ... and {len(format_info.product_types) - 20} more")
    else:
        print("Usage: python dynamic_template_analyzer.py <template_file.xlsm>")

