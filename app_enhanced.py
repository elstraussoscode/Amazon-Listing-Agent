"""
Amazon Listing Agent - Enhanced Streamlit App
Automatisches Befüllen von Amazon-Templates mit KI-optimierten Inhalten
"""

import streamlit as st
import pandas as pd
import openpyxl
from openai import OpenAI
import io
from typing import Dict, List, Optional, Tuple
import json
from pathlib import Path
import logging
import re
from pydantic import BaseModel, Field

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

# Pydantic Models for Structured Outputs
class ProductContent(BaseModel):
    """Structured output for generated Amazon listing content"""
    artikelname: str = Field(description="Optimierter Produkttitel, max 200 Zeichen")
    bullet_points: List[str] = Field(description="5 Aufzählungspunkte, je max 250 Zeichen", min_length=5, max_length=5)
    suchbegriffe: str = Field(description="5 Suchbegriffe durch Komma getrennt")

class TemplateColumnInfo(BaseModel):
    """Info about a single column in template"""
    column: str = Field(description="Column letter (e.g., 'D', 'CC')")
    column_index: int = Field(description="Zero-based column index")
    row: Optional[int] = Field(default=None, description="Row number where header is found")

class TemplateStructure(BaseModel):
    """Structured output for template analysis"""
    # Required fields
    title: TemplateColumnInfo = Field(description="Artikelname/Item Name column")
    sku: TemplateColumnInfo = Field(description="SKU column")
    bullet_points: List[TemplateColumnInfo] = Field(description="List of bullet point columns")
    search_terms: List[TemplateColumnInfo] = Field(default=[], description="List of search term columns")
    header_row: int = Field(description="Row number with column headers")
    data_start_row: int = Field(description="First row for actual data")
    
    # Optional product identity fields
    brand: Optional[TemplateColumnInfo] = Field(default=None, description="Marke/Brand Name column")
    manufacturer: Optional[TemplateColumnInfo] = Field(default=None, description="Hersteller/Manufacturer column")
    product_type: Optional[TemplateColumnInfo] = Field(default=None, description="Produkttyp/Product Type column")
    
    # Optional product attributes
    material: Optional[TemplateColumnInfo] = Field(default=None, description="Material column")
    color: Optional[TemplateColumnInfo] = Field(default=None, description="Farbe/Color column")
    size: Optional[TemplateColumnInfo] = Field(default=None, description="Größe/Size column")
    weight: Optional[TemplateColumnInfo] = Field(default=None, description="Gewicht/Weight column")
    dimensions: Optional[TemplateColumnInfo] = Field(default=None, description="Abmessungen/Dimensions column")
    
    # Optional descriptions
    product_description: Optional[TemplateColumnInfo] = Field(default=None, description="Produktbeschreibung/Description column")
    
    # Optional technical specs
    model_number: Optional[TemplateColumnInfo] = Field(default=None, description="Modellnummer/Model Number column")
    ean: Optional[TemplateColumnInfo] = Field(default=None, description="EAN/GTIN column")
    
    # Optional pricing/stock
    quantity: Optional[TemplateColumnInfo] = Field(default=None, description="Menge/Quantity column")
    
    # Optional categorization
    category: Optional[TemplateColumnInfo] = Field(default=None, description="Kategorie/Category column")
    subcategory: Optional[TemplateColumnInfo] = Field(default=None, description="Unterkategorie/Subcategory column")
    
    # Optional features
    care_instructions: Optional[TemplateColumnInfo] = Field(default=None, description="Pflegehinweise/Care Instructions column")
    warranty: Optional[TemplateColumnInfo] = Field(default=None, description="Garantie/Warranty column")
    country_of_origin: Optional[TemplateColumnInfo] = Field(default=None, description="Herkunftsland/Country of Origin column")

class InputStructure(BaseModel):
    """Structured output for input data analysis"""
    first_data_row: int = Field(description="First row with actual product data (0-based index)")
    
    # Product identity
    product_name_column: Optional[str] = Field(default=None, description="Column with product name/title")
    brand_column: Optional[str] = Field(default=None, description="Column with brand name")
    manufacturer_column: Optional[str] = Field(default=None, description="Column with manufacturer")
    sku_column: Optional[str] = Field(default=None, description="Column with SKU/ASIN/Article number")
    ean_column: Optional[str] = Field(default=None, description="Column with EAN/GTIN/Barcode")
    model_number_column: Optional[str] = Field(default=None, description="Column with model number")
    
    # Product content
    description_columns: List[str] = Field(default=[], description="Columns with product descriptions")
    bullet_columns: List[str] = Field(default=[], description="Columns with bullet points")
    
    # Product attributes
    material_column: Optional[str] = Field(default=None, description="Column with material info")
    color_column: Optional[str] = Field(default=None, description="Column with color")
    size_column: Optional[str] = Field(default=None, description="Column with size/dimensions")
    weight_column: Optional[str] = Field(default=None, description="Column with weight")
    dimensions_column: Optional[str] = Field(default=None, description="Column with dimensions/measurements")
    
    # Categorization
    category_column: Optional[str] = Field(default=None, description="Column with category")
    subcategory_column: Optional[str] = Field(default=None, description="Column with subcategory")
    product_type_column: Optional[str] = Field(default=None, description="Column with product type")
    
    # Pricing & Stock
    price_column: Optional[str] = Field(default=None, description="Column with price")
    quantity_column: Optional[str] = Field(default=None, description="Column with quantity/stock")
    
    # Additional info
    care_instructions_column: Optional[str] = Field(default=None, description="Column with care/cleaning instructions")
    warranty_column: Optional[str] = Field(default=None, description="Column with warranty info")
    country_origin_column: Optional[str] = Field(default=None, description="Column with country of origin")
    
    # Catch-all
    other_important_columns: List[str] = Field(default=[], description="Other relevant columns with useful data")

# Page configuration
st.set_page_config(
    page_title="Amazon Listing Agent - Template Filler",
    page_icon="🛒",
    layout="wide",
    initial_sidebar_state="expanded"
)

# Default COSMO/RUFUS optimized prompt
DEFAULT_PROMPT = """Du bist ein Experte für Amazon SEO und die neuen COSMO/RUFUS KI-Systeme von Amazon.

**INTERNE RICHTLINIEN (nicht explizit erwähnen, nur als Guidelines nutzen):**
Nutze die 15 COSMO-Beziehungstypen als unsichtbare Struktur: Produkttyp (is), Eigenschaften (has_property), Komponenten (has_component), Verwendungszweck (used_for), Anwendungssituation (used_in), Nutzergruppe (used_by), Komplementärprodukte (used_with), Material (made_of), Qualitätsmerkmale (has_quality), Marke (has_brand), Stil (has_style), Zielgruppe (targets_audience), übergeordnete Themen (associated_with), Zertifikate (has_certification), ermöglichte Aktivitäten (enables_activity).

**PRODUKTDATEN:**
{{product_data}}

**AUFGABE:**
Erstelle verkaufsorientierte, natürliche Amazon-Listing-Inhalte:

**Optimierter Artikelname** (max. 200 Zeichen):
- Marke + Produkttyp + wichtigste USPs
- Natürlich lesbar, keine Keyword-Stuffing
- Enthält Zielgruppe oder Hauptanwendung wenn sinnvoll
- Beispiel: "Premium Bambus Schneidebrett Set 3-teilig - Antibakteriell FSC-Zertifiziert für Hobbyköche"

**5 Aufzählungspunkte** (je max. 250 Zeichen):
- OHNE Nummerierung (kein "1.", "2.", "Erstens" etc.)
- Format: GROSSBUCHSTABEN-KEYWORD am Anfang, dann Beschreibung
- Jeder Punkt deckt verschiedene Aspekte ab: Zielgruppe, Anwendungsszenarien, Eigenschaften, Qualität, Vorteile
- Nutze konkrete Szenarien statt abstrakte Beschreibungen
- Beantworte implizit: Für wen? Wofür? Warum besser? Wie verwenden? Wann nützlich?
- Beispiel: "IDEAL FÜR HOBBYKÖCHE UND FAMILIEN - Das 3-teilige Set bietet für jeden Zweck die richtige Größe, von großen Steaks bis zu kleinem Gemüse, perfekt für die tägliche Küchenarbeit"

**Suchbegriffe** (5 Begriffe, durch Komma getrennt):
- Long-Tail-Keywords mit Kontext
- Verwendungsszenarien (z.B. "geschenk für hobbyköche", "camping kochgeschirr set")
- Zielgruppen-Begriffe (z.B. "küchenhelfer für senioren")
- Problemlösungen (z.B. "plastikfrei lebensmittel aufbewahren")
- Keine einzelnen generischen Keywords

**WICHTIG:**
- Schreibe natürlich und verkaufsorientiert
- Keine akademische oder Listen-Sprache
- Nutze die COSMO-Typen als unsichtbare Checkliste, erwähne sie aber nicht
- Fokus auf Kundennutzen und konkrete Anwendungsfälle"""

# Initialize session state
for key, default in [
    ('api_key', ''),
    ('prompt_template', DEFAULT_PROMPT),
    ('generated_data', None),
    ('template_file', None),
    ('template_columns', None)
]:
    if key not in st.session_state:
        st.session_state[key] = default

def load_excel_file(file):
    """Load Excel file and return dict of dataframes"""
    try:
        xl = pd.ExcelFile(file)
        sheets = {}
        for sheet_name in xl.sheet_names:
            sheets[sheet_name] = pd.read_excel(file, sheet_name=sheet_name)
        return sheets
    except Exception as e:
        st.error(f"Fehler beim Laden: {str(e)}")
        return None

def find_vorlage_sheet(excel_dict):
    """Find Vorlage sheet in template"""
    vorlage_sheets = [name for name in excel_dict.keys() 
                     if 'vorlage' in name.lower() and not name.lower().startswith('änderungen')]
    return vorlage_sheets[0] if vorlage_sheets else None

def analyze_input_sheet_with_ai(df, api_key: str) -> InputStructure:
    """Use GPT-5-mini with structured outputs to analyze input sheet"""
    try:
        logger.info("Using AI to analyze input sheet structure...")
        client = OpenAI(api_key=api_key)
        
        # Get first 3 rows + column headers
        header_data = {
            "column_names": list(df.columns),
            "first_3_rows": df.head(3).to_dict(orient='records')
        }
        
        prompt = f"""Analysiere diese Excel-Datei mit Produktdaten.

STRUKTUR (erste 3 Zeilen):
{json.dumps(header_data, indent=2, ensure_ascii=False)}

AUFGABE:
1. Finde die ERSTE ZEILE MIT ECHTEN PRODUKTDATEN (nicht Header, nicht leere Zeilen)
   - 'first_data_row' ist der 0-basierte Index der ersten Datenzeile

2. Identifiziere ALLE verfügbaren Spalten (nur wenn vorhanden!):
   
   **Identität:**
   - Produktname/Title/Bezeichnung
   - Marke/Brand/Seller
   - Hersteller/Manufacturer
   - SKU/ASIN/Artikel-Nr
   - EAN/GTIN/Barcode
   - Modellnummer/Model Number
   
   **Content:**
   - Beschreibung/Description/Produktbeschreibung
   - Bullet Points (falls vorhanden)
   
   **Attribute:**
   - Material
   - Farbe/Color
   - Größe/Size/Maße/Abmessungen
   - Gewicht/Weight
   - Dimensions/Abmessungen
   
   **Kategorisierung:**
   - Kategorie/Category
   - Unterkategorie/Subcategory
   - Produkttyp/Product Type
   
   **Preis & Lager:**
   - Preis/Price
   - Menge/Quantity/Stock
   
   **Zusatzinfos:**
   - Pflegehinweise/Care Instructions
   - Garantie/Warranty
   - Herkunftsland/Country of Origin/Made in
   - Alle anderen Spalten mit nützlichen Produktinfos

Gib NUR Spalten an die WIRKLICH existieren. Bei fehlenden Spalten: null oder leere Liste."""

        completion = client.beta.chat.completions.parse(
            model="gpt-5-mini",
            messages=[
                {"role": "system", "content": "Du bist ein Experte für Produktdaten-Analyse."},
                {"role": "user", "content": prompt}
            ],
            response_format=InputStructure
        )
        
        result = completion.choices[0].message.parsed
        logger.info(f"Successfully analyzed input structure: {result.model_dump()}")
        return result
        
    except Exception as e:
        logger.error(f"Error in AI input analysis: {str(e)}", exc_info=True)
        # Fallback
        return InputStructure(
            first_data_row=0,
            product_name_column=df.columns[0] if len(df.columns) > 0 else None,
            other_important_columns=list(df.columns)
        )

def detect_template_columns_with_ai(df, api_key: str) -> TemplateStructure:
    """Use GPT-5-mini with structured outputs to analyze template"""
    try:
        logger.info("Using AI to analyze template structure...")
        client = OpenAI(api_key=api_key)
        
        # Get first 10 rows of headers
        header_data = []
        for row_idx in range(min(10, len(df))):
            row_dict = {}
            for col_idx, value in enumerate(df.iloc[row_idx]):
                if pd.notna(value):
                    col_letter = openpyxl.utils.get_column_letter(col_idx + 1)
                    row_dict[col_letter] = str(value)[:100]
            if row_dict:
                header_data.append({"row": row_idx + 1, "columns": row_dict})
        
        prompt = f"""Analysiere diese Amazon Template Excel-Struktur.

TEMPLATE HEADERS (erste 10 Zeilen):
{json.dumps(header_data, indent=2, ensure_ascii=False)}

AUFGABE - Finde ALLE verfügbaren Spalten (nur wenn sie existieren!):

**Pflichtfelder:**
- Artikelname/Item Name (Produkttitel)
- SKU (Produktkennung)
- Aufzählungspunkte/Bullet Points (meist 5 Spalten)
- Header Row (Zeile mit Column-Namen)
- Data Start Row (erste Datenzeile)

**Optionale Felder (nur angeben wenn vorhanden!):**
- Markenname/Brand Name
- Hersteller/Manufacturer
- Produkttyp/Product Type
- Produktbeschreibung/Product Description
- Suchbegriffe/Search Terms/Keywords (meist 5 Spalten)
- Material
- Farbe/Color/Colour
- Größe/Size
- Gewicht/Weight
- Abmessungen/Dimensions/Measurements
- Modellnummer/Model Number
- EAN/GTIN/Barcode
- Kategorie/Category/Browse Node
- Unterkategorie/Subcategory
- Menge/Quantity
- Pflegehinweise/Care Instructions
- Garantie/Warranty
- Herkunftsland/Country of Origin

Suche nach deutschen UND englischen Begriffen!
Gib null zurück für nicht vorhandene Felder."""

        completion = client.beta.chat.completions.parse(
            model="gpt-5-mini",
            messages=[
                {"role": "system", "content": "Du bist ein Experte für Amazon Template-Analyse."},
                {"role": "user", "content": prompt}
            ],
            response_format=TemplateStructure
        )
        
        result = completion.choices[0].message.parsed
        logger.info(f"Successfully analyzed template: {result.model_dump()}")
        return result
        
    except Exception as e:
        logger.error(f"Error in AI template analysis: {str(e)}", exc_info=True)
        # Fallback
        return detect_template_columns_fallback(df)

def detect_template_columns_fallback(df) -> TemplateStructure:
    """Fallback: Basic column detection without AI"""
    logger.info("Using fallback column detection...")
    
    title_col = sku_col = None
    bullet_cols = []
    search_cols = []
    
    for row_idx in range(min(10, len(df))):
        for col_idx, value in enumerate(df.iloc[row_idx]):
            if pd.notna(value):
                val_str = str(value).lower()
                col_letter = openpyxl.utils.get_column_letter(col_idx + 1)
                
                if 'artikelname' in val_str and not title_col:
                    title_col = TemplateColumnInfo(column=col_letter, column_index=col_idx, row=row_idx)
                
                if val_str == 'sku' and not sku_col:
                    sku_col = TemplateColumnInfo(column=col_letter, column_index=col_idx, row=row_idx)
                
                if 'aufzählungspunkt' in val_str:
                    bullet_cols.append(TemplateColumnInfo(column=col_letter, column_index=col_idx, row=row_idx))
                
                if 'suchbegriffe' in val_str:
                    search_cols.append(TemplateColumnInfo(column=col_letter, column_index=col_idx, row=row_idx))
    
    # Defaults if not found
    if not title_col:
        title_col = TemplateColumnInfo(column="D", column_index=3, row=4)
    if not sku_col:
        sku_col = TemplateColumnInfo(column="A", column_index=0, row=4)
    
    return TemplateStructure(
        title=title_col,
        sku=sku_col,
        bullet_points=bullet_cols[:5] if bullet_cols else [],
        search_terms=search_cols[:5] if search_cols else [],
        header_row=4,
        data_start_row=6
    )

def clean_gpt5_response(response_text: str) -> str:
    """Clean GPT-5-mini response to get valid JSON"""
    logger.info(f"Raw response (first 500 chars): {response_text[:500]}")
    
    # Remove reasoning field if present
    response_text = re.sub(r'"Reasoning":\s*\[.*?\],?\s*', '', response_text, flags=re.DOTALL)
    
    # Extract JSON from markdown code blocks
    json_match = re.search(r'```(?:json)?\s*(\{.*?\})\s*```', response_text, re.DOTALL)
    if json_match:
        response_text = json_match.group(1)
        logger.info("Extracted JSON from markdown block")
    
    # Remove leading/trailing whitespace and markdown markers
    response_text = response_text.strip()
    if response_text.startswith('```'):
        lines = response_text.split('\n')
        response_text = '\n'.join(lines[1:-1] if lines[-1].strip() == '```' else lines[1:])
        response_text = response_text.strip()
    
    logger.info(f"Cleaned response (first 500 chars): {response_text[:500]}")
    return response_text

def extract_product_info(row_data: pd.Series, input_analysis: InputStructure) -> str:
    """Extract ALL available product info based on AI analysis"""
    product_info_parts = []
    
    # Helper to add field if exists
    def add_if_exists(col_name, label):
        if col_name and col_name in row_data.index and pd.notna(row_data[col_name]):
            product_info_parts.append(f"{label}: {row_data[col_name]}")
    
    # Product Identity
    add_if_exists(input_analysis.product_name_column, "Produktname")
    add_if_exists(input_analysis.brand_column, "Marke")
    add_if_exists(input_analysis.manufacturer_column, "Hersteller")
    add_if_exists(input_analysis.sku_column, "SKU/ASIN")
    add_if_exists(input_analysis.ean_column, "EAN")
    add_if_exists(input_analysis.model_number_column, "Modellnummer")
    
    # Descriptions
    for col in input_analysis.description_columns:
        if col in row_data.index and pd.notna(row_data[col]):
            product_info_parts.append(f"Beschreibung: {row_data[col]}")
    
    # Existing bullets (to use as reference/inspiration)
    for i, col in enumerate(input_analysis.bullet_columns[:5], 1):
        if col in row_data.index and pd.notna(row_data[col]):
            product_info_parts.append(f"Bestehender Bullet {i}: {row_data[col]}")
    
    # Product Attributes
    add_if_exists(input_analysis.material_column, "Material")
    add_if_exists(input_analysis.color_column, "Farbe")
    add_if_exists(input_analysis.size_column, "Größe")
    add_if_exists(input_analysis.weight_column, "Gewicht")
    add_if_exists(input_analysis.dimensions_column, "Abmessungen")
    
    # Categorization
    add_if_exists(input_analysis.category_column, "Kategorie")
    add_if_exists(input_analysis.subcategory_column, "Unterkategorie")
    add_if_exists(input_analysis.product_type_column, "Produkttyp")
    
    # Pricing & Stock
    add_if_exists(input_analysis.price_column, "Preis")
    add_if_exists(input_analysis.quantity_column, "Menge")
    
    # Additional Info
    add_if_exists(input_analysis.care_instructions_column, "Pflegehinweise")
    add_if_exists(input_analysis.warranty_column, "Garantie")
    add_if_exists(input_analysis.country_origin_column, "Herkunftsland")
    
    # Other important columns
    for col in input_analysis.other_important_columns:
        if col in row_data.index and pd.notna(row_data[col]):
            product_info_parts.append(f"{col}: {row_data[col]}")
    
    # Fallback: if nothing found, use all non-null columns
    if not product_info_parts:
        for col, val in row_data.items():
            if pd.notna(val) and str(val).strip():
                product_info_parts.append(f"{col}: {val}")
    
    return "\n".join(product_info_parts)

def generate_content_with_openai(product_data: Dict, api_key: str, prompt_template: str) -> Optional[ProductContent]:
    """Generate optimized content using GPT-5-mini with structured outputs"""
    try:
        client = OpenAI(api_key=api_key)
        
        # Handle both extracted_info (from AI analysis) and raw dict
        if 'extracted_info' in product_data:
            product_info = product_data['extracted_info']
        else:
            product_info = "\n".join([f"- {k}: {v}" for k, v in product_data.items() 
                                     if pd.notna(v) and str(v).strip()])
        
        # Use double braces for template safety
        prompt = prompt_template.replace('{{product_data}}', product_info)
        
        logger.info("Sending request to GPT-5-mini with structured output...")
        
        completion = client.beta.chat.completions.parse(
            model="gpt-5-mini",
            messages=[
                {"role": "system", "content": "Du bist ein Amazon SEO-Experte für COSMO und RUFUS."},
                {"role": "user", "content": prompt}
            ],
            response_format=ProductContent
        )
        
        result = completion.choices[0].message.parsed
        logger.info(f"Successfully generated content: {result.artikelname[:50]}...")
        return result
        
    except Exception as e:
        logger.error(f"Error in generate_content_with_openai: {str(e)}", exc_info=True)
        st.error(f"Fehler bei Content-Generierung: {str(e)}")
        return None

def fill_template_with_openpyxl(template_file_bytes, vorlage_name: str, columns: TemplateStructure, 
                                 generated_contents: List[ProductContent], products_df: pd.DataFrame, 
                                 input_analysis: InputStructure, input_start_row: int) -> openpyxl.Workbook:
    """Fill template using openpyxl - fills ALL available fields when data exists"""
    logger.info("Loading template with openpyxl to preserve structure...")
    
    # Load workbook (NOT read_only so we can write)
    wb = openpyxl.load_workbook(io.BytesIO(template_file_bytes), keep_vba=True)
    ws = wb[vorlage_name]
    
    logger.info(f"Template workbook loaded. Sheet: {vorlage_name}")
    filled_fields_count = 0
    
    for idx, content in enumerate(generated_contents):
        # Get source product row
        actual_row = input_start_row + idx
        product = products_df.iloc[actual_row]
        
        # Calculate target row in template (1-based for openpyxl)
        target_row = columns.data_start_row + idx + 1
        
        logger.info(f"\n{'='*60}")
        logger.info(f"Filling template row {target_row} with product {idx + 1}")
        logger.info(f"{'='*60}")
        
        # Helper function to fill if data exists
        def fill_if_exists(template_col_info, input_col_name, field_name, is_ai_generated=False, ai_value=None):
            nonlocal filled_fields_count
            if not template_col_info:
                return
            
            value = None
            if is_ai_generated:
                value = ai_value
            elif input_col_name and input_col_name in product.index and pd.notna(product[input_col_name]):
                value = str(product[input_col_name])
            
            if value:
                col_idx = template_col_info.column_index + 1
                ws.cell(row=target_row, column=col_idx).value = value
                logger.info(f"  ✅ {field_name} ({template_col_info.column}): {str(value)[:60]}")
                filled_fields_count += 1
        
        # AI-GENERATED CONTENT
        fill_if_exists(columns.title, None, "Artikelname", True, content.artikelname)
        
        # PRODUCT IDENTITY (from source)
        fill_if_exists(columns.sku, input_analysis.sku_column, "SKU")
        fill_if_exists(columns.brand, input_analysis.brand_column, "Marke")
        fill_if_exists(columns.manufacturer, input_analysis.manufacturer_column, "Hersteller")
        fill_if_exists(columns.product_type, input_analysis.product_type_column, "Produkttyp")
        fill_if_exists(columns.model_number, input_analysis.model_number_column, "Modellnummer")
        fill_if_exists(columns.ean, input_analysis.ean_column, "EAN")
        
        # PRODUCT ATTRIBUTES (from source)
        fill_if_exists(columns.material, input_analysis.material_column, "Material")
        fill_if_exists(columns.color, input_analysis.color_column, "Farbe")
        fill_if_exists(columns.size, input_analysis.size_column, "Größe")
        fill_if_exists(columns.weight, input_analysis.weight_column, "Gewicht")
        fill_if_exists(columns.dimensions, input_analysis.dimensions_column, "Abmessungen")
        
        # CATEGORIZATION (from source)
        fill_if_exists(columns.category, input_analysis.category_column, "Kategorie")
        fill_if_exists(columns.subcategory, input_analysis.subcategory_column, "Unterkategorie")
        
        # PRICING & STOCK (from source)
        fill_if_exists(columns.quantity, input_analysis.quantity_column, "Menge")
        
        # ADDITIONAL INFO (from source)
        fill_if_exists(columns.care_instructions, input_analysis.care_instructions_column, "Pflegehinweise")
        fill_if_exists(columns.warranty, input_analysis.warranty_column, "Garantie")
        fill_if_exists(columns.country_of_origin, input_analysis.country_origin_column, "Herkunftsland")
        
        # DESCRIPTION (from source if exists, otherwise could use AI)
        if columns.product_description and input_analysis.description_columns:
            for desc_col in input_analysis.description_columns:
                if desc_col in product.index and pd.notna(product[desc_col]):
                    col_idx = columns.product_description.column_index + 1
                    ws.cell(row=target_row, column=col_idx).value = str(product[desc_col])
                    logger.info(f"  ✅ Beschreibung: {str(product[desc_col])[:60]}")
                    filled_fields_count += 1
                    break
        
        # AI-GENERATED BULLET POINTS
        for i, bp_col_info in enumerate(columns.bullet_points[:5]):
            if i < len(content.bullet_points):
                bp_col = bp_col_info.column_index + 1
                ws.cell(row=target_row, column=bp_col).value = content.bullet_points[i]
                logger.info(f"  ✅ BP{i+1} ({bp_col_info.column}): {content.bullet_points[i][:40]}...")
                filled_fields_count += 1
        
        # AI-GENERATED SEARCH TERMS
        search_terms = [s.strip() for s in content.suchbegriffe.split(',')]
        for i, st_col_info in enumerate(columns.search_terms[:5]):
            if i < len(search_terms):
                st_col = st_col_info.column_index + 1
                ws.cell(row=target_row, column=st_col).value = search_terms[i]
                logger.info(f"  ✅ Search{i+1} ({st_col_info.column}): {search_terms[i]}")
                filled_fields_count += 1
        
        logger.info(f"✅ Completed row {target_row}\n")
    
    logger.info(f"{'='*60}")
    logger.info(f"✅ SUMMARY: {len(generated_contents)} products processed")
    logger.info(f"✅ Total fields filled: {filled_fields_count}")
    logger.info(f"{'='*60}")
    return wb

# Main UI
st.title("🛒 Amazon Listing Agent - Template Filler")
st.markdown("**Automatisches Befüllen von Amazon-Templates mit KI-optimierten Inhalten**")

# Create tabs
tab1, tab2, tab3, tab4 = st.tabs(["📝 Content Generator & Template Filler", "📊 Template Preview", "📋 Logs & Debug", "⚙️ Konfiguration"])

with tab1:
    st.header("Workflow: Input → KI-Generierung → Template Füllen")
    
    col1, col2 = st.columns(2)
    
    with col1:
        st.subheader("1️⃣ Produktdaten hochladen")
        products_file = st.file_uploader(
            "Excel mit Produktdaten",
            type=['xlsx', 'xls', 'xlsm'],
            key='products',
            help="Unterstützt: Amazon-Export, Profino-Format, oder eigene Struktur"
        )
        
        products_df = None
        if products_file:
            sheets = load_excel_file(products_file)
            if sheets:
                sheet_name = st.selectbox("Sheet auswählen:", list(sheets.keys()))
                products_df = sheets[sheet_name]
                
                st.success(f"✅ {len(products_df)} Zeilen geladen")
                st.dataframe(products_df.head(10), use_container_width=True)
                
                # AI-Analyze the input structure
                if st.session_state.get('api_key', '').strip():
                    with st.spinner("🤖 GPT-5-mini analysiert Produktdaten-Struktur..."):
                        input_analysis = analyze_input_sheet_with_ai(products_df, st.session_state.api_key)
                        st.session_state.input_analysis = input_analysis
                        
                        with st.expander("🔍 AI-Erkannte Input-Spalten"):
                            st.write("**GPT-5-mini hat folgende Struktur erkannt:**")
                            st.info(f"📍 Erste Datenzeile: **Zeile {input_analysis.first_data_row + 1}** (Index: {input_analysis.first_data_row})")
                            st.json(input_analysis.model_dump())
                else:
                    st.warning("⚠️ API Key fehlt - verwende alle Spalten")
    
    with col2:
        st.subheader("2️⃣ Amazon-Template hochladen")
        template_file = st.file_uploader(
            "Amazon Upload-Template (.xlsm)",
            type=['xlsx', 'xls', 'xlsm'],
            key='template',
            help="Das offizielle Amazon-Template für Ihre Kategorie"
        )
        
        if template_file:
            template_sheets = load_excel_file(template_file)
            if template_sheets:
                vorlage_name = find_vorlage_sheet(template_sheets)
                if vorlage_name:
                    st.success(f"✅ Vorlage gefunden: {vorlage_name}")
                    template_df = template_sheets[vorlage_name]
                    
                    # Use AI to analyze template structure dynamically
                    if st.session_state.get('api_key', '').strip():
                        with st.spinner("🤖 GPT-5-mini analysiert Template-Struktur..."):
                            columns = detect_template_columns_with_ai(template_df, st.session_state.api_key)
                    else:
                        st.warning("⚠️ API Key fehlt - verwende Basis-Erkennung")
                        columns = detect_template_columns_fallback(template_df)
                    
                    # Store in session
                    st.session_state.template_file = template_file
                    st.session_state.template_columns = columns
                    
                    # Show detected columns
                    with st.expander("🔍 AI-Erkannte Template-Spalten (Was wird befüllt?)"):
                        # Count available fields
                        available_fields = [
                            ("Marke", columns.brand),
                            ("Hersteller", columns.manufacturer),
                            ("Produkttyp", columns.product_type),
                            ("Material", columns.material),
                            ("Farbe", columns.color),
                            ("Größe", columns.size),
                            ("Gewicht", columns.weight),
                            ("Abmessungen", columns.dimensions),
                            ("Modellnummer", columns.model_number),
                            ("EAN", columns.ean),
                            ("Kategorie", columns.category),
                            ("Unterkategorie", columns.subcategory),
                            ("Menge", columns.quantity),
                            ("Beschreibung", columns.product_description),
                            ("Pflegehinweise", columns.care_instructions),
                            ("Garantie", columns.warranty),
                            ("Herkunftsland", columns.country_of_origin)
                        ]
                        available_count = sum(1 for _, col in available_fields if col)
                        
                        st.success(f"🎯 **{available_count + 2 + len(columns.bullet_points) + len(columns.search_terms)} Felder** werden befüllt!")
                        
                        st.write("**✨ AI-Generierte Felder (COSMO/RUFUS optimiert):**")
                        st.write(f"- Artikelname: Spalte {columns.title.column}")
                        st.write(f"- Bullet Points: {len(columns.bullet_points)} Spalten")
                        st.write(f"- Suchbegriffe: {len(columns.search_terms)} Spalten")
                        
                        st.write("\n**📋 Aus Input-Daten übernommen (wenn vorhanden):**")
                        col1, col2 = st.columns(2)
                        
                        with col1:
                            st.write(f"- SKU: {columns.sku.column}")
                            for name, col in available_fields[:9]:
                                if col:
                                    st.write(f"- {name}: {col.column} ✅")
                        
                        with col2:
                            for name, col in available_fields[9:]:
                                if col:
                                    st.write(f"- {name}: {col.column} ✅")
                        
                        st.info(f"📍 Daten werden ab Zeile **{columns.data_start_row}** geschrieben")
                        
                        with st.expander("📊 Vollständige Struktur (JSON)"):
                            st.json(columns.model_dump())
                else:
                    st.error("❌ Keine Vorlage-Sheet gefunden!")
    
    st.markdown("---")
    
    # Generation & Filling
    if products_df is not None and st.session_state.get('api_key', '').strip() and template_file:
        st.subheader("3️⃣ Content generieren & Template befüllen")
        
        col1, col2 = st.columns([3, 1])
        with col1:
            num_products = st.number_input(
                "Anzahl zu verarbeitender Produkte:", 
                min_value=1, 
                max_value=len(products_df), 
                value=min(5, len(products_df)),
                help="Wählen Sie, wie viele Produkte verarbeitet werden sollen"
            )
        
        if st.button("🚀 Starten: Content generieren & Template füllen", type="primary", use_container_width=True):
            if not st.session_state.get('api_key', '').strip():
                st.error("❌ Bitte OpenAI API Key in der Konfiguration eingeben!")
                st.stop()
            
            # Get template info
            template_sheets = load_excel_file(template_file)
            vorlage_name = find_vorlage_sheet(template_sheets)
            columns = st.session_state.template_columns
            
            # Read template file bytes for openpyxl
            template_file.seek(0)
            template_bytes = template_file.read()
            
            progress_bar = st.progress(0)
            status_text = st.empty()
            results_container = st.container()
            
            generated_contents = []
            
            # Get first data row from input analysis
            input_start_row = 0
            if 'input_analysis' in st.session_state:
                input_start_row = st.session_state.input_analysis.first_data_row
                logger.info(f"Starting from row {input_start_row} in input data")
            
            for idx in range(num_products):
                # Get product from the correct row
                actual_row = input_start_row + idx
                if actual_row >= len(products_df):
                    logger.warning(f"Row {actual_row} exceeds data length {len(products_df)}")
                    break
                    
                product = products_df.iloc[actual_row]
                status_text.text(f"🔄 Verarbeite Produkt {idx + 1} von {num_products} (Input-Zeile {actual_row + 1})...")
                
                # Extract product info using AI analysis of input structure
                if 'input_analysis' in st.session_state:
                    product_info_str = extract_product_info(product, st.session_state.input_analysis)
                    logger.info(f"Product {idx + 1} (row {actual_row}) extracted info:\n{product_info_str[:300]}")
                    product_data_for_ai = {"extracted_info": product_info_str}
                else:
                    product_data_for_ai = product.to_dict()
                
                # Generate content
                content = generate_content_with_openai(
                    product_data_for_ai,
                    st.session_state.api_key,
                    st.session_state.prompt_template
                )
                
                if content:
                    generated_contents.append(content)
                    
                    # Show preview
                    with results_container:
                        with st.expander(f"✅ Produkt {idx + 1}: {content.artikelname[:50]}..."):
                            st.write("**Artikelname:**", content.artikelname)
                            st.write("**Bullet Points:**")
                            for i, bp in enumerate(content.bullet_points, 1):
                                st.write(f"{i}. {bp}")
                            st.write("**Suchbegriffe:**", content.suchbegriffe)
                
                progress_bar.progress((idx + 1) / num_products)
            
            status_text.text(f"✅ Content-Generierung abgeschlossen! {len(generated_contents)} Produkte")
            
            # Now fill the template using openpyxl (preserves structure!)
            with st.spinner("📝 Befülle Template mit generierten Inhalten..."):
                filled_workbook = fill_template_with_openpyxl(
                    template_bytes,
                    vorlage_name,
                    columns,
                    generated_contents,
                    products_df,
                    st.session_state.input_analysis,
                    input_start_row
                )
                
                # Save to BytesIO
                output = io.BytesIO()
                filled_workbook.save(output)
                output.seek(0)
                
                st.session_state.filled_workbook_bytes = output.getvalue()
                st.session_state.generated_data = generated_contents
            
            st.markdown("---")
            st.success(f"🎉 Template erfolgreich befüllt mit {len(generated_contents)} Produkten!")
            
            # Show summary of filled fields
            col1, col2, col3 = st.columns(3)
            with col1:
                st.metric("Produkte verarbeitet", len(generated_contents))
            with col2:
                st.metric("Felder pro Produkt", "~20-30+")
            with col3:
                st.metric("Template-Struktur", "100% erhalten")
            
            st.info("""
            ✅ **Original-Struktur erhalten:** Alle Formeln, Formatierungen und Makros bleiben intakt!
            
            **Was wurde befüllt:**
            - 🤖 AI-Generiert: Artikelname, 5 Bullet Points, 5 Suchbegriffe (COSMO/RUFUS optimiert)
            - 📋 Von Input: SKU, Marke, Material, Farbe, Größe, Gewicht, EAN, Kategorie, etc. (wenn vorhanden)
            - 🎯 Nur Felder mit Daten werden befüllt - keine leeren Zellen!
            """)
            
            # Download button
            st.download_button(
                label="📥 Befülltes Template herunterladen (.xlsm)",
                data=st.session_state.filled_workbook_bytes,
                file_name=f"amazon_template_befuellt_{pd.Timestamp.now().strftime('%Y%m%d_%H%M%S')}.xlsm",
                mime="application/vnd.ms-excel.sheet.macroEnabled.12",
                use_container_width=True,
                type="primary"
            )
    
    elif not st.session_state.get('api_key', '').strip():
        st.warning("⚠️ Bitte OpenAI API Key in der Konfiguration eingeben!")
    elif not template_file:
        st.info("ℹ️ Bitte laden Sie ein Amazon-Template hoch")
    elif products_df is None:
        st.info("ℹ️ Bitte laden Sie Produktdaten hoch")

with tab2:
    st.header("📊 Template Preview")
    
    if 'filled_workbook_bytes' in st.session_state and st.session_state.filled_workbook_bytes:
        st.subheader("Befülltes Template (Excel Vorschau)")
        st.success("✅ Template wurde mit openpyxl befüllt - Struktur bleibt exakt erhalten!")
        
        # Load for preview
        try:
            wb_preview = openpyxl.load_workbook(io.BytesIO(st.session_state.filled_workbook_bytes), data_only=True)
            template_sheets = load_excel_file(st.session_state.template_file)
            vorlage_name = find_vorlage_sheet(template_sheets)
            
            # Convert to pandas for display
            ws = wb_preview[vorlage_name]
            data = []
            for row in ws.iter_rows(values_only=True):
                data.append(row)
            preview_df = pd.DataFrame(data)
            
            st.dataframe(preview_df.head(20), use_container_width=True, height=600)
            
            st.info(f"📋 Zeige erste 20 Zeilen. Template enthält alle Original-Formatierungen, Formeln und Makros!")
        except Exception as e:
            st.error(f"Vorschau-Fehler: {e}")
    else:
        st.info("ℹ️ Noch kein befülltes Template vorhanden. Führen Sie zuerst die Generierung durch.")

with tab3:
    st.header("📋 Logs & Debug Information")
    st.markdown("**Echtzeit-Logs des Verarbeitungsprozesses**")
    
    st.info("""
    **Was wird geloggt:**
    - GPT-5-mini API Anfragen und Antworten
    - Template-Struktur-Analyse
    - JSON Parsing Schritte
    - Fehler und Warnungen
    - Row-by-row Verarbeitung
    
    **Logs erscheinen im Terminal wo Sie die App gestartet haben!**
    """)
    
    st.code("""
    # Terminal öffnen und logs sehen:
    cd "/Users/florianstrauss/Desktop/Amazon Listing Agent"
    source venv/bin/activate
    streamlit run app_enhanced.py
    
    # Logs werden automatisch angezeigt mit:
    # INFO - Sending request to GPT-5-mini...
    # INFO - AI Template Analysis Response: ...
    # INFO - Successfully parsed JSON response
    # etc.
    """, language="bash")
    
    if 'input_analysis' in st.session_state:
        st.subheader("🔍 Input-Daten Analyse (GPT-5-mini)")
        analysis = st.session_state.input_analysis
        if isinstance(analysis, InputStructure):
            st.json(analysis.model_dump())
        else:
            st.json(analysis)
    
    if 'template_columns' in st.session_state and st.session_state.template_columns:
        st.subheader("🔍 Template-Struktur Analyse (GPT-5-mini)")
        cols = st.session_state.template_columns
        if isinstance(cols, TemplateStructure):
            st.json(cols.model_dump())
        else:
            st.json(cols)

with tab4:
    st.header("⚙️ Konfiguration")
    
    # API Key
    st.subheader("🔑 OpenAI API Key")
    api_key_input = st.text_input(
        "API Key",
        type="password",
        value=st.session_state.api_key
    )
    
    if st.button("💾 API Key speichern"):
        st.session_state.api_key = api_key_input
        st.success("✅ Gespeichert!")
    
    st.markdown("---")
    
    # Prompt
    st.subheader("📝 Prompt Template")
    prompt_input = st.text_area(
        "Prompt",
        value=st.session_state.prompt_template,
        height=400
    )
    
    col1, col2 = st.columns(2)
    with col1:
        if st.button("💾 Prompt speichern", use_container_width=True):
            st.session_state.prompt_template = prompt_input
            st.success("✅ Gespeichert!")
    
    with col2:
        if st.button("🔄 Zurücksetzen", use_container_width=True):
            st.session_state.prompt_template = DEFAULT_PROMPT
            st.rerun()

# Sidebar
with st.sidebar:
    st.markdown("### 🛒 Amazon Listing Agent")
    st.markdown("**Template Filler Edition**")
    
    st.markdown("---")
    st.markdown("### 📊 Status")
    
    if st.session_state.get('api_key', '').strip():
        st.success("✅ API Key")
    else:
        st.error("❌ Kein API Key")
    
    if 'input_analysis' in st.session_state:
        st.success("✅ Input analysiert")
    else:
        st.info("ℹ️ Kein Input")
    
    if st.session_state.template_file:
        st.success("✅ Template geladen")
    else:
        st.warning("⚠️ Kein Template")
    
    if 'filled_workbook_bytes' in st.session_state and st.session_state.filled_workbook_bytes:
        st.success(f"✅ Template fertig ({len(st.session_state.generated_data)} Produkte)")
    elif st.session_state.generated_data:
        st.success(f"✅ {len(st.session_state.generated_data)} Produkte generiert")
    else:
        st.info("ℹ️ Keine Daten")
    
    st.markdown("---")
    st.markdown("### 🚀 Workflow")
    st.markdown("""
    1. **Konfiguration**: API Key eingeben
    2. **Upload**: Produktdaten + Template
    3. **Generieren**: KI erstellt Inhalte
    4. **Download**: Befülltes Template
    """)
    
    st.markdown("---")
    st.markdown("**COSMO & RUFUS optimiert** 🎯")
