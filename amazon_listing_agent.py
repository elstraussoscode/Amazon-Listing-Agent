"""
Amazon Listing Agent - Excel Upload Sheet Processor
Automatically fills in Amazon seller and vendor upload sheets
"""

import tkinter as tk
from tkinter import filedialog, messagebox, ttk, scrolledtext
import pandas as pd
import openpyxl
from pathlib import Path
from typing import Dict, List, Tuple, Optional
import os


class AmazonListingAgent:
    def __init__(self, root):
        self.root = root
        self.root.title("Amazon Listing Agent")
        self.root.geometry("1000x700")
        
        self.current_file = None
        self.workbook = None
        self.vorlage_sheet = None
        self.column_mapping = {}
        self.header_row = None
        
        self.setup_ui()
    
    def setup_ui(self):
        """Setup the user interface"""
        # Main container
        main_frame = ttk.Frame(self.root, padding="10")
        main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        # Configure grid weights
        self.root.columnconfigure(0, weight=1)
        self.root.rowconfigure(0, weight=1)
        main_frame.columnconfigure(0, weight=1)
        main_frame.rowconfigure(4, weight=1)
        
        # Title
        title_label = ttk.Label(main_frame, text="Amazon Listing Agent", 
                               font=('Helvetica', 16, 'bold'))
        title_label.grid(row=0, column=0, pady=10, sticky=tk.W)
        
        # File selection section
        file_frame = ttk.LabelFrame(main_frame, text="1. Upload Sheet auswählen", padding="10")
        file_frame.grid(row=1, column=0, sticky=(tk.W, tk.E), pady=5)
        file_frame.columnconfigure(1, weight=1)
        
        ttk.Button(file_frame, text="Datei hochladen", 
                  command=self.upload_file).grid(row=0, column=0, padx=5)
        
        self.file_label = ttk.Label(file_frame, text="Keine Datei ausgewählt", 
                                     foreground="gray")
        self.file_label.grid(row=0, column=1, sticky=tk.W, padx=10)
        
        # Column detection info
        self.info_frame = ttk.LabelFrame(main_frame, text="2. Erkannte Spalten", padding="10")
        self.info_frame.grid(row=2, column=0, sticky=(tk.W, tk.E), pady=5)
        self.info_frame.columnconfigure(0, weight=1)
        
        self.info_text = scrolledtext.ScrolledText(self.info_frame, height=8, width=80)
        self.info_text.grid(row=0, column=0, sticky=(tk.W, tk.E))
        self.info_text.insert('1.0', "Bitte laden Sie zuerst eine Excel-Datei hoch...")
        self.info_text.config(state='disabled')
        
        # Data entry section
        data_frame = ttk.LabelFrame(main_frame, text="3. Daten eingeben", padding="10")
        data_frame.grid(row=3, column=0, sticky=(tk.W, tk.E), pady=5)
        data_frame.columnconfigure(1, weight=1)
        
        # Row selector
        ttk.Label(data_frame, text="Zeile:").grid(row=0, column=0, sticky=tk.W, pady=5)
        self.row_spinbox = ttk.Spinbox(data_frame, from_=1, to=100, width=10)
        self.row_spinbox.grid(row=0, column=1, sticky=tk.W, padx=5)
        self.row_spinbox.set(1)
        
        ttk.Button(data_frame, text="Zeile laden", 
                  command=self.load_row_data).grid(row=0, column=2, padx=5)
        
        # Identifiers section
        ids_frame = ttk.LabelFrame(data_frame, text="Identifier (für Updates)", padding="5")
        ids_frame.grid(row=1, column=0, columnspan=3, sticky=(tk.W, tk.E), pady=10)
        ids_frame.columnconfigure(1, weight=1)
        ids_frame.columnconfigure(3, weight=1)
        
        ttk.Label(ids_frame, text="SKU:").grid(row=0, column=0, sticky=tk.W, padx=5)
        self.sku_entry = ttk.Entry(ids_frame, width=30)
        self.sku_entry.grid(row=0, column=1, sticky=(tk.W, tk.E), padx=5)
        
        ttk.Label(ids_frame, text="ASIN:").grid(row=0, column=2, sticky=tk.W, padx=5)
        self.asin_entry = ttk.Entry(ids_frame, width=30)
        self.asin_entry.grid(row=0, column=3, sticky=(tk.W, tk.E), padx=5)
        
        ttk.Label(ids_frame, text="Angebotsaktion:").grid(row=1, column=0, sticky=tk.W, padx=5)
        self.action_combo = ttk.Combobox(ids_frame, width=28, 
                                         values=["Partial Update", "Update", "Delete"])
        self.action_combo.grid(row=1, column=1, sticky=(tk.W, tk.E), padx=5)
        self.action_combo.set("Partial Update")
        
        # Main data entry
        ttk.Label(data_frame, text="Artikelname (Titel):").grid(row=2, column=0, 
                                                                sticky=tk.W, pady=5)
        self.title_entry = ttk.Entry(data_frame, width=80)
        self.title_entry.grid(row=2, column=1, columnspan=2, sticky=(tk.W, tk.E), padx=5)
        
        # Bullet points
        bp_frame = ttk.LabelFrame(data_frame, text="Aufzählungspunkte (Bullet Points)", 
                                  padding="5")
        bp_frame.grid(row=3, column=0, columnspan=3, sticky=(tk.W, tk.E), pady=10)
        bp_frame.columnconfigure(1, weight=1)
        
        self.bullet_entries = []
        for i in range(5):
            ttk.Label(bp_frame, text=f"BP {i+1}:").grid(row=i, column=0, 
                                                        sticky=tk.W, padx=5, pady=2)
            entry = ttk.Entry(bp_frame, width=80)
            entry.grid(row=i, column=1, sticky=(tk.W, tk.E), padx=5, pady=2)
            self.bullet_entries.append(entry)
        
        # Search terms
        ttk.Label(data_frame, text="Suchbegriffe:").grid(row=4, column=0, 
                                                         sticky=tk.W, pady=5)
        self.search_terms_entry = ttk.Entry(data_frame, width=80)
        self.search_terms_entry.grid(row=4, column=1, columnspan=2, 
                                     sticky=(tk.W, tk.E), padx=5)
        
        # Action buttons
        button_frame = ttk.Frame(main_frame)
        button_frame.grid(row=5, column=0, pady=10)
        
        ttk.Button(button_frame, text="In Excel schreiben", 
                  command=self.write_to_excel).grid(row=0, column=0, padx=5)
        
        ttk.Button(button_frame, text="Excel speichern", 
                  command=self.save_excel).grid(row=0, column=1, padx=5)
        
        ttk.Button(button_frame, text="Felder leeren", 
                  command=self.clear_fields).grid(row=0, column=2, padx=5)
        
        # Status bar
        self.status_label = ttk.Label(main_frame, text="Bereit", 
                                      relief=tk.SUNKEN, anchor=tk.W)
        self.status_label.grid(row=6, column=0, sticky=(tk.W, tk.E), pady=5)
    
    def upload_file(self):
        """Handle file upload"""
        file_path = filedialog.askopenfilename(
            title="Excel-Datei auswählen",
            filetypes=[("Excel files", "*.xlsx;*.xlsm;*.xls"), ("All files", "*.*")]
        )
        
        if not file_path:
            return
        
        try:
            self.current_file = file_path
            self.workbook = openpyxl.load_workbook(file_path)
            
            # Find the Vorlage sheet
            vorlage_sheets = [s for s in self.workbook.sheetnames 
                            if 'vorlage' in s.lower()]
            
            if not vorlage_sheets:
                messagebox.showerror("Fehler", 
                    "Keine 'Vorlage'-Sheet gefunden!\n" + 
                    f"Verfügbare Sheets: {', '.join(self.workbook.sheetnames)}")
                return
            
            self.vorlage_sheet = self.workbook[vorlage_sheets[0]]
            self.file_label.config(text=Path(file_path).name, foreground="green")
            
            # Detect columns
            self.detect_columns()
            self.status_label.config(text=f"Datei geladen: {Path(file_path).name}")
            
        except Exception as e:
            messagebox.showerror("Fehler", f"Fehler beim Laden der Datei:\n{str(e)}")
            self.status_label.config(text="Fehler beim Laden")
    
    def detect_columns(self):
        """Detect column positions by header names"""
        if not self.vorlage_sheet:
            return
        
        self.column_mapping = {}
        self.header_row = None
        
        # Search for headers in first 10 rows
        key_terms = {
            'artikelname': 'title',
            'aufzählungspunkt': 'bullet_points',
            'suchbegriffe': 'search_terms',
            'sku': 'sku',
            'asin': 'asin',
            'angebotsaktion': 'action'
        }
        
        bullet_columns = []
        
        for row_idx in range(1, 11):
            for cell in self.vorlage_sheet[row_idx]:
                if cell.value:
                    cell_str = str(cell.value).lower()
                    
                    for term, key in key_terms.items():
                        if term in cell_str:
                            if term == 'aufzählungspunkt':
                                bullet_columns.append(cell.column)
                                if self.header_row is None:
                                    self.header_row = row_idx
                            else:
                                self.column_mapping[key] = cell.column
                                if self.header_row is None:
                                    self.header_row = row_idx
        
        # Store bullet point columns
        if bullet_columns:
            self.column_mapping['bullet_points'] = sorted(bullet_columns)
        
        # Display detected columns
        self.display_column_info()
    
    def display_column_info(self):
        """Display detected column information"""
        self.info_text.config(state='normal')
        self.info_text.delete('1.0', tk.END)
        
        info = f"Sheet: {self.vorlage_sheet.title}\n"
        info += f"Header-Zeile: {self.header_row}\n"
        info += f"Daten beginnen bei Zeile: {self.header_row + 1 if self.header_row else 'unbekannt'}\n\n"
        info += "Erkannte Spalten:\n"
        info += "-" * 60 + "\n"
        
        column_names = {
            'title': 'Artikelname',
            'bullet_points': 'Aufzählungspunkte',
            'search_terms': 'Suchbegriffe',
            'sku': 'SKU',
            'asin': 'ASIN',
            'action': 'Angebotsaktion'
        }
        
        for key, display_name in column_names.items():
            if key in self.column_mapping:
                col_val = self.column_mapping[key]
                if key == 'bullet_points':
                    col_letters = [openpyxl.utils.get_column_letter(c) for c in col_val]
                    info += f"  {display_name}: {', '.join(col_letters)}\n"
                else:
                    col_letter = openpyxl.utils.get_column_letter(col_val)
                    info += f"  {display_name}: {col_letter}\n"
            else:
                info += f"  {display_name}: nicht gefunden\n"
        
        self.info_text.insert('1.0', info)
        self.info_text.config(state='disabled')
    
    def load_row_data(self):
        """Load existing data from selected row"""
        if not self.vorlage_sheet or not self.header_row:
            messagebox.showwarning("Warnung", "Bitte laden Sie zuerst eine Datei!")
            return
        
        try:
            row_num = int(self.row_spinbox.get())
            data_row = self.header_row + row_num
            
            # Load title
            if 'title' in self.column_mapping:
                cell_value = self.vorlage_sheet.cell(data_row, 
                                                     self.column_mapping['title']).value
                self.title_entry.delete(0, tk.END)
                if cell_value:
                    self.title_entry.insert(0, str(cell_value))
            
            # Load bullet points
            if 'bullet_points' in self.column_mapping:
                for i, col in enumerate(self.column_mapping['bullet_points'][:5]):
                    cell_value = self.vorlage_sheet.cell(data_row, col).value
                    self.bullet_entries[i].delete(0, tk.END)
                    if cell_value:
                        self.bullet_entries[i].insert(0, str(cell_value))
            
            # Load search terms
            if 'search_terms' in self.column_mapping:
                cell_value = self.vorlage_sheet.cell(data_row, 
                                                     self.column_mapping['search_terms']).value
                self.search_terms_entry.delete(0, tk.END)
                if cell_value:
                    self.search_terms_entry.insert(0, str(cell_value))
            
            # Load SKU
            if 'sku' in self.column_mapping:
                cell_value = self.vorlage_sheet.cell(data_row, 
                                                     self.column_mapping['sku']).value
                self.sku_entry.delete(0, tk.END)
                if cell_value:
                    self.sku_entry.insert(0, str(cell_value))
            
            # Load ASIN
            if 'asin' in self.column_mapping:
                cell_value = self.vorlage_sheet.cell(data_row, 
                                                     self.column_mapping['asin']).value
                self.asin_entry.delete(0, tk.END)
                if cell_value:
                    self.asin_entry.insert(0, str(cell_value))
            
            # Load Action
            if 'action' in self.column_mapping:
                cell_value = self.vorlage_sheet.cell(data_row, 
                                                     self.column_mapping['action']).value
                if cell_value:
                    self.action_combo.set(str(cell_value))
            
            self.status_label.config(text=f"Zeile {row_num} geladen")
            
        except ValueError:
            messagebox.showerror("Fehler", "Bitte geben Sie eine gültige Zeilennummer ein!")
        except Exception as e:
            messagebox.showerror("Fehler", f"Fehler beim Laden der Zeile:\n{str(e)}")
    
    def write_to_excel(self):
        """Write data to Excel sheet"""
        if not self.vorlage_sheet or not self.header_row:
            messagebox.showwarning("Warnung", "Bitte laden Sie zuerst eine Datei!")
            return
        
        try:
            row_num = int(self.row_spinbox.get())
            data_row = self.header_row + row_num
            
            # Write title
            if 'title' in self.column_mapping:
                self.vorlage_sheet.cell(data_row, 
                                       self.column_mapping['title']).value = self.title_entry.get()
            
            # Write bullet points
            if 'bullet_points' in self.column_mapping:
                for i, col in enumerate(self.column_mapping['bullet_points'][:5]):
                    self.vorlage_sheet.cell(data_row, col).value = self.bullet_entries[i].get()
            
            # Write search terms
            if 'search_terms' in self.column_mapping:
                self.vorlage_sheet.cell(data_row, 
                                       self.column_mapping['search_terms']).value = \
                    self.search_terms_entry.get()
            
            # Write SKU
            if 'sku' in self.column_mapping and self.sku_entry.get():
                self.vorlage_sheet.cell(data_row, 
                                       self.column_mapping['sku']).value = self.sku_entry.get()
            
            # Write ASIN
            if 'asin' in self.column_mapping and self.asin_entry.get():
                self.vorlage_sheet.cell(data_row, 
                                       self.column_mapping['asin']).value = self.asin_entry.get()
            
            # Write Action
            if 'action' in self.column_mapping:
                self.vorlage_sheet.cell(data_row, 
                                       self.column_mapping['action']).value = self.action_combo.get()
            
            self.status_label.config(text=f"Daten in Zeile {row_num} geschrieben")
            messagebox.showinfo("Erfolg", 
                f"Daten erfolgreich in Zeile {row_num} geschrieben!\n\n" + 
                "Vergessen Sie nicht, die Datei zu speichern.")
            
        except ValueError:
            messagebox.showerror("Fehler", "Bitte geben Sie eine gültige Zeilennummer ein!")
        except Exception as e:
            messagebox.showerror("Fehler", f"Fehler beim Schreiben:\n{str(e)}")
    
    def save_excel(self):
        """Save the Excel file"""
        if not self.workbook or not self.current_file:
            messagebox.showwarning("Warnung", "Keine Datei zum Speichern!")
            return
        
        try:
            # Ask for save location
            save_path = filedialog.asksaveasfilename(
                title="Excel-Datei speichern",
                defaultextension=".xlsx",
                initialfile=Path(self.current_file).stem + "_bearbeitet.xlsx",
                filetypes=[("Excel files", "*.xlsx"), ("Excel macro files", "*.xlsm")]
            )
            
            if not save_path:
                return
            
            self.workbook.save(save_path)
            self.status_label.config(text=f"Datei gespeichert: {Path(save_path).name}")
            messagebox.showinfo("Erfolg", f"Datei erfolgreich gespeichert:\n{save_path}")
            
        except Exception as e:
            messagebox.showerror("Fehler", f"Fehler beim Speichern:\n{str(e)}")
    
    def clear_fields(self):
        """Clear all input fields"""
        self.title_entry.delete(0, tk.END)
        for entry in self.bullet_entries:
            entry.delete(0, tk.END)
        self.search_terms_entry.delete(0, tk.END)
        self.sku_entry.delete(0, tk.END)
        self.asin_entry.delete(0, tk.END)
        self.action_combo.set("Partial Update")
        self.status_label.config(text="Felder geleert")


def main():
    root = tk.Tk()
    app = AmazonListingAgent(root)
    root.mainloop()


if __name__ == "__main__":
    main()
